"""
scraper.py — Baixa as planilhas de ITBI da Prefeitura de SP e salva no banco SQLite.
Execute com:  python scraper.py
              python scraper.py --anos 2024 2025 2026
              python scraper.py --forcar
"""

import re
import hashlib
import requests
import pandas as pd
import db as _db
from datetime import datetime
from pathlib import Path

# ──────────────────────────────────────────────
DB_PATH   = Path(__file__).parent / "itbi.db"
CACHE_DIR = Path(__file__).parent / "cache"
CACHE_DIR.mkdir(exist_ok=True)

ARQUIVOS = {
    2026: "https://prefeitura.sp.gov.br/documents/d/fazenda/guias-de-itbi-pagas-2-xlsx",
    2025: "https://prefeitura.sp.gov.br/cidade/secretarias/upload/fazenda/arquivos/itbi/GUIAS%20DE%20ITBI%20PAGAS%20%2828012026%29%20XLS.xlsx",
    2024: "https://prefeitura.sp.gov.br/cidade/secretarias/upload/fazenda/arquivos/itbi/GUIAS-DE-ITBI-PAGAS-2024.xlsx",
    2023: "https://www.prefeitura.sp.gov.br/cidade/secretarias/upload/fazenda/arquivos/XLSX/GUIAS-DE-ITBI-PAGAS-2023.xlsx",
    2022: "https://www.prefeitura.sp.gov.br/cidade/secretarias/upload/fazenda/arquivos/XLSX/GUIAS_DE_ITBI_PAGAS_12-2022.xlsx",
    2021: "https://www.prefeitura.sp.gov.br/cidade/secretarias/upload/fazenda/arquivos/itbi/ITBI_Setembro_2022/GUIAS_DE_ITBI_PAGAS_(2021).xlsx",
    2020: "https://www.prefeitura.sp.gov.br/cidade/secretarias/upload/fazenda/arquivos/itbi/ITBI_Setembro_2022/GUIAS_DE_ITBI_PAGAS_(2020).xlsx",
    2019: "https://www.prefeitura.sp.gov.br/cidade/secretarias/upload/fazenda/arquivos/itbi/ITBI_Setembro_2022/GUIAS_DE_ITBI_PAGAS_(2019).xlsx",
    2018: "https://www.prefeitura.sp.gov.br/cidade/secretarias/upload/fazenda/arquivos/itbi/guias_de_itbi_pagas_2018.xlsx",
    2017: "https://www.prefeitura.sp.gov.br/cidade/secretarias/upload/fazenda/arquivos/itbi/guias_de_itbi_pagas_2017.xlsx",
    2016: "https://www.prefeitura.sp.gov.br/cidade/secretarias/upload/fazenda/arquivos/itbi/guias_de_itbi_pagas_2016.xlsx",
    2015: "https://www.prefeitura.sp.gov.br/cidade/secretarias/upload/fazenda/arquivos/itbi/guias_de_itbi_pagas_2015.xlsx",
    2014: "https://www.prefeitura.sp.gov.br/cidade/secretarias/upload/fazenda/arquivos/itbi/guias_de_itbi_pagas_2014.xlsx",
    2013: "https://www.prefeitura.sp.gov.br/cidade/secretarias/upload/fazenda/arquivos/itbi/guias_de_itbi_pagas_2013.xlsx",
    2012: "https://www.prefeitura.sp.gov.br/cidade/secretarias/upload/fazenda/arquivos/itbi/guias_de_itbi_pagas_2012.xlsx",
    2011: "https://www.prefeitura.sp.gov.br/cidade/secretarias/upload/fazenda/arquivos/itbi/guias_de_itbi_pagas_2011.xlsx",
    2010: "https://www.prefeitura.sp.gov.br/cidade/secretarias/upload/fazenda/arquivos/itbi/guias_de_itbi_pagas_2010.xlsx",
    2009: "https://www.prefeitura.sp.gov.br/cidade/secretarias/upload/fazenda/arquivos/itbi/guias_de_itbi_pagas_2009.xlsx",
    2008: "https://www.prefeitura.sp.gov.br/cidade/secretarias/upload/fazenda/arquivos/itbi/guias_de_itbi_pagas_2008.xlsx",
    2007: "https://www.prefeitura.sp.gov.br/cidade/secretarias/upload/fazenda/arquivos/itbi/guias_de_itbi_pagas_2007.xlsx",
    2006: "https://www.prefeitura.sp.gov.br/cidade/secretarias/upload/fazenda/arquivos/itbi/guias_de_itbi_pagas_2006.xlsx",
}

HEADERS = {"User-Agent": "Mozilla/5.0 (compatible; ITBI-Dashboard/1.0)"}

MESES_PT = {
    "jan": 1, "fev": 2, "mar": 3, "abr": 4,
    "mai": 5, "jun": 6, "jul": 7, "ago": 8,
    "set": 9, "out": 10, "nov": 11, "dez": 12,
}


# ──────────────────────────────────────────────
# BANCO DE DADOS
# ──────────────────────────────────────────────

def init_db():
    _db.init_db()
    return

def _init_db_old():  # kept for reference
    import sqlite3
    conn = sqlite3.connect("itbi.db")
    conn.execute("""
        CREATE TABLE IF NOT EXISTS transacoes (
            id                    INTEGER PRIMARY KEY AUTOINCREMENT,
            ano_referencia        INTEGER,
            mes_referencia        INTEGER,
            data_transacao        TEXT,
            logradouro            TEXT,
            numero                TEXT,
            complemento           TEXT,
            bairro                TEXT,
            cep                   TEXT,
            sql_terreno           TEXT,
            area_terreno          REAL,
            area_construida       REAL,
            valor_declarado       REAL,
            valor_financiado      REAL,
            valor_itbi            REAL,
            natureza_transacao    TEXT,
            tipo_uso              TEXT,
            descricao_uso         TEXT,
            proporcao_transmitida REAL,
            tipo_financiamento    TEXT,
            acc_iptu              TEXT,
            valor_venal_ref       REAL,
            cartorio_registro     TEXT,
            matricula_imovel      TEXT,
            situacao_sql          TEXT,
            testada               REAL,
            fracao_ideal          TEXT,
            padrao_iptu           TEXT,
            created_at            TEXT DEFAULT (datetime('now'))
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS arquivos_processados (
            ano        INTEGER PRIMARY KEY,
            hash       TEXT,
            linhas     INTEGER,
            updated_at TEXT
        )
    """)
    # Migração automática
    novas_colunas = [
        ("descricao_uso",        "TEXT"),
        ("proporcao_transmitida", "REAL"),
        ("tipo_financiamento",    "TEXT"),
        ("acc_iptu",              "TEXT"),
        ("valor_venal_ref",       "REAL"),
        ("cartorio_registro",     "TEXT"),
        ("matricula_imovel",      "TEXT"),
        ("situacao_sql",          "TEXT"),
        ("testada",               "REAL"),
        ("fracao_ideal",          "TEXT"),
        ("padrao_iptu",           "TEXT"),
    ]
    cols_existentes = {row[1] for row in conn.execute("PRAGMA table_info(transacoes)")}
    for nome, tipo in novas_colunas:
        if nome not in cols_existentes:
            conn.execute(f"ALTER TABLE transacoes ADD COLUMN {nome} {tipo}")
            print(f"   🔧 Coluna added: {nome}")

    conn.execute("CREATE INDEX IF NOT EXISTS idx_logradouro ON transacoes(logradouro COLLATE NOCASE)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_bairro     ON transacoes(bairro COLLATE NOCASE)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_ano        ON transacoes(ano_referencia)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_mes        ON transacoes(mes_referencia)")
    conn.commit()
    conn.close()
    print("✅ Banco de dados pronto:", DB_PATH)


# ──────────────────────────────────────────────
# DOWNLOAD
# ──────────────────────────────────────────────

def file_hash(path: Path) -> str:
    h = hashlib.md5()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


def download(ano: int, url: str) -> Path | None:
    cache_path = CACHE_DIR / f"itbi_{ano}.xlsx"
    print(f"📥 [{ano}] Baixando...", end=" ", flush=True)
    try:
        r = requests.get(url, headers=HEADERS, timeout=120, stream=True)
        r.raise_for_status()
        with open(cache_path, "wb") as f:
            for chunk in r.iter_content(chunk_size=32768):
                f.write(chunk)
        print(f"✅ ({cache_path.stat().st_size / 1024:.0f} KB)")
        return cache_path
    except Exception as e:
        print(f"❌ {e}")
        return None


# ──────────────────────────────────────────────
# NORMALIZAÇÃO DE COLUNAS
# ──────────────────────────────────────────────

def limpar_col(col: str) -> str:
    """Remove acentos e normaliza nome de coluna."""
    col = str(col).lower().strip()
    col = col.translate(str.maketrans(
        'áàãâäéèêëíìîïóòõôöúùûüçñ',
        'aaaaaeeeeiiiiooooouuuucn'
    ))
    col = re.sub(r'[^\w]', '_', col)
    col = re.sub(r'_+', '_', col)
    return col.strip('_')


# Mapeamento baseado nos nomes REAIS das colunas (verificados nas planilhas)
COL_MAP = {
    # SQL
    "n_do_cadastro_sql_":                               "sql_terreno",
    "n_do_cadastro__sql_":                              "sql_terreno",
    "n_cadastro_sql":                                   "sql_terreno",
    "sql":                                              "sql_terreno",
    "nro_contribuinte":                                 "sql_terreno",

    # Logradouro
    "nome_do_logradouro":                               "logradouro",
    "logradouro":                                       "logradouro",
    "nome_logradouro":                                  "logradouro",
    "endereco":                                         "logradouro",

    # Número
    "numero":                                           "numero",
    "num":                                              "numero",

    # Complemento
    "complemento":                                      "complemento",

    # Bairro
    "bairro":                                           "bairro",

    # CEP
    "cep":                                              "cep",

    # Natureza
    "natureza_de_transacao":                            "natureza_transacao",
    "natureza_transacao":                               "natureza_transacao",
    "natureza_da_transacao":                            "natureza_transacao",
    "natureza":                                         "natureza_transacao",

    # Valor declarado (coluna longa — captura variações de tamanho)
    "valor_de_transacao_declarado_pelo_contribuinte_":  "valor_declarado",
    "valor_de_transacao_declarado_pelo_contribuinte":   "valor_declarado",
    "valor_de_transacao__declarado_pelo_contribuinte_": "valor_declarado",
    "valor_declarado":                                  "valor_declarado",
    "valor_transacao":                                  "valor_declarado",
    "vl_declarado":                                     "valor_declarado",

    # Base de cálculo ITBI
    "base_de_calculo_adotada":                          "valor_itbi",
    "base_calculo_adotada":                             "valor_itbi",
    "valor_venal_de_referencia_proporcional_":          "valor_itbi",
    "valor_venal_de_referencia_proporcional":           "valor_itbi",
    "valor_itbi":                                       "valor_itbi",
    "vl_itbi":                                          "valor_itbi",

    # Proporção transmitida
    "proporcao_transmitida____":                        "proporcao_transmitida",
    "proporcao_transmitida___":                         "proporcao_transmitida",
    "proporcao_transmitida__":                          "proporcao_transmitida",
    "proporcao_transmitida_":                           "proporcao_transmitida",
    "proporcao_transmitida":                            "proporcao_transmitida",

    # Valor financiado
    "valor_financiado":                                 "valor_financiado",
    "vl_financiado":                                    "valor_financiado",

    # Data de transação
    "data_de_transacao":                                "data_transacao",
    "data_transacao":                                   "data_transacao",
    "dt_transacao":                                     "data_transacao",

    # Área terreno
    "area_do_terreno_m2_":                              "area_terreno",
    "area_do_terreno__m2_":                             "area_terreno",
    "area_terreno":                                     "area_terreno",

    # Área construída
    "area_construida_m2_":                              "area_construida",
    "area_construida__m2_":                             "area_construida",
    "area_construida":                                  "area_construida",

    # Uso
    "uso_iptu_":                                        "tipo_uso",
    "uso__iptu_":                                       "tipo_uso",
    "uso_iptu":                                         "tipo_uso",
    "tipo_uso":                                         "tipo_uso",
    "uso":                                              "tipo_uso",

    "descricao_do_uso_iptu_":                           "descricao_uso",
    "descricao_do_uso__iptu_":                          "descricao_uso",
    "descricao_do_uso_iptu":                            "descricao_uso",

    # Tipo de financiamento
    "tipo_de_financiamento":                            "tipo_financiamento",
    "tipo_financiamento":                               "tipo_financiamento",

    # ACC IPTU
    "acc_iptu_":                                        "acc_iptu",
    "acc__iptu_":                                       "acc_iptu",
    "acc_iptu":                                         "acc_iptu",
    "acc":                                              "acc_iptu",
    # Valor venal de referência (cheio)
    "valor_venal_de_referencia":                        "valor_venal_ref",
    "valor_venal_referencia":                           "valor_venal_ref",
    # Cartório e matrícula
    "cartorio_de_registro":                             "cartorio_registro",
    "cartorio_registro":                                "cartorio_registro",
    "matricula_do_imovel":                              "matricula_imovel",
    "matricula_imovel":                                 "matricula_imovel",
    # Situação SQL
    "situacao_do_sql":                                  "situacao_sql",
    "situacao_sql":                                     "situacao_sql",
    # Testada e fração
    "testada_m_":                                       "testada",
    "testada__m_":                                      "testada",
    "testada":                                          "testada",
    "fracao_ideal":                                     "fracao_ideal",
    # Padrão IPTU
    "padrao_iptu_":                                     "padrao_iptu",
    "padrao__iptu_":                                    "padrao_iptu",
    "padrao_iptu":                                      "padrao_iptu",
    "descricao_do_padrao_iptu_":                        "padrao_iptu",
}

COLUNAS_FINAIS = [
    "ano_referencia", "mes_referencia", "data_transacao",
    "logradouro", "numero", "complemento", "bairro", "cep",
    "sql_terreno", "area_terreno", "area_construida",
    "valor_declarado", "valor_financiado", "valor_itbi",
    "valor_venal_ref", "proporcao_transmitida",
    "natureza_transacao", "tipo_financiamento",
    "tipo_uso", "descricao_uso", "acc_iptu",
    "cartorio_registro", "matricula_imovel", "situacao_sql",
    "testada", "fracao_ideal", "padrao_iptu",
]


def mes_da_aba(nome_aba: str) -> int | None:
    nome = nome_aba.lower().strip()
    for sigla, num in MESES_PT.items():
        if nome.startswith(sigla):
            return num
    return None


def normalizar_df(df: pd.DataFrame, ano: int, mes: int | None) -> pd.DataFrame:
    # Limpa e normaliza nomes de colunas
    df.columns = [limpar_col(c) for c in df.columns]
    df = df.loc[:, ~df.columns.duplicated()]

    # Tenta capturar colunas de "valor declarado" com nome truncado
    # (a prefeitura às vezes corta o nome longo da coluna)
    for col in df.columns:
        if col not in COL_MAP:
            if col.startswith("valor_de_transacao"):
                COL_MAP[col] = "valor_declarado"
            elif col.startswith("proporcao_transmitida"):
                COL_MAP[col] = "proporcao_transmitida"
            elif col.startswith("area_do_terreno"):
                COL_MAP[col] = "area_terreno"
            elif col.startswith("area_construida"):
                COL_MAP[col] = "area_construida"
            elif col.startswith("base_de_calculo"):
                COL_MAP[col] = "valor_itbi"
            elif col.startswith("valor_venal_de_referencia_prop"):
                COL_MAP[col] = "valor_itbi"
            elif col.startswith("descricao_do_uso"):
                COL_MAP[col] = "descricao_uso"
            elif col.startswith("uso_") or col == "uso":
                COL_MAP[col] = "tipo_uso"
            elif col.startswith("tipo_de_financiamento") or col.startswith("tipo_financiamento"):
                COL_MAP[col] = "tipo_financiamento"
            elif col.startswith("acc_") or col == "acc":
                COL_MAP[col] = "acc_iptu"

    rename = {c: COL_MAP[c] for c in df.columns if c in COL_MAP}
    df = df.rename(columns=rename)

    # Deduplica novamente após o rename — duas colunas diferentes podem
    # ter mapeado para o mesmo nome (ex: base_calculo e valor_venal → valor_itbi)
    df = df.loc[:, ~df.columns.duplicated(keep='first')]

    df["ano_referencia"] = ano
    df["mes_referencia"] = mes

    # Converte numéricos (sem encadear .str para evitar problemas com pandas 3.x)
    def limpar_numero(x):
        if x is None:
            return None
        s = str(x).strip()
        # Remove espaços e símbolos monetários
        s = re.sub(r'[\s	]', '', s)
        s = re.sub(r'[R$]', '', s)
        s = s.strip()
        if not s:
            return None

        tem_virgula = ',' in s
        num_pontos  = s.count('.')

        if tem_virgula and num_pontos >= 1:
            # Formato BR completo: 1.700.000,25 ou 542.156,25
            s = s.replace('.', '').replace(',', '.')
        elif tem_virgula and num_pontos == 0:
            # Formato BR sem milhar: 542156,25 ou 542,25
            s = s.replace(',', '.')
        elif not tem_virgula and num_pontos == 1:
            # Float internacional do openpyxl: 542156.25 — NÃO remove o ponto
            pass
        elif not tem_virgula and num_pontos >= 2:
            # Separadores de milhar sem decimal: 1.700.000
            s = s.replace('.', '')
        # else: inteiro puro sem ponto/vírgula — deixa como está

        # Remove qualquer caractere não numérico restante (exceto ponto e sinal)
        s = re.sub(r'[^\d.\-]', '', s)
        return s if s else None

    for col in ["valor_declarado", "valor_financiado", "valor_itbi",
                "area_terreno", "area_construida", "proporcao_transmitida"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col].apply(limpar_numero), errors="coerce")

    # Garante todas as colunas finais
    for col in COLUNAS_FINAIS:
        if col not in df.columns:
            df[col] = None

    # Remove linhas totalmente vazias (cabeçalhos duplicados, etc.)
    df = df[df["logradouro"].notna() | df["bairro"].notna() | df["valor_declarado"].notna()]
    df = df[df["logradouro"] != "Nome do Logradouro"]  # remove linha de cabeçalho duplicado

    return df[COLUNAS_FINAIS]


# ──────────────────────────────────────────────
# PROCESSAR XLSX — TODOS OS 12 MESES
# ──────────────────────────────────────────────

def processar_xlsx(path: Path, ano: int) -> pd.DataFrame:
    import math
    import traceback
    from openpyxl import load_workbook

    # Abre UMA vez sem read_only (mais compatível com arquivos complexos)
    wb = load_workbook(path, data_only=True)
    frames = []

    IGNORAR = {"legenda", "sumario", "sumário", "instrucao", "leia",
               "orientacao", "explicac"}

    for sheet_name in wb.sheetnames:
        if any(x in sheet_name.lower() for x in IGNORAR):
            continue

        mes = mes_da_aba(sheet_name)
        try:
            ws = wb[sheet_name]

            # Lê todas as linhas como tuplas de valores
            todas_linhas = []
            for row in ws.iter_rows(values_only=True):
                todas_linhas.append(row)

            if len(todas_linhas) < 2:
                continue

            # Encontra o header real (linha com pelo menos 2 palavras-chave)
            PALAVRAS = {"logradouro", "bairro", "cadastro", "natureza",
                        "valor", "cep", "numero", "complemento"}
            header_idx = 0
            for i, row in enumerate(todas_linhas[:15]):
                vals = {str(v).lower() for v in row if v is not None}
                if len(vals & PALAVRAS) >= 2:
                    header_idx = i
                    break

            raw_headers = todas_linhas[header_idx]
            data_rows   = todas_linhas[header_idx + 1:]

            if not data_rows:
                continue

            # Cria nomes de coluna únicos
            headers = []
            seen: dict = {}
            for j, h in enumerate(raw_headers):
                name = str(h).strip() if h is not None else f"_col{j}"
                if name in seen:
                    seen[name] += 1
                    name = f"{name}_{seen[name]}"
                else:
                    seen[name] = 0
                headers.append(name)

            # Converte valores para string ou None
            def to_str(v):
                if v is None:
                    return None
                if isinstance(v, float) and math.isnan(v):
                    return None
                s = str(v).strip()
                return None if s in ("", "None", "nan", "NaN") else s

            # Monta registros alinhando colunas
            records = []
            n = len(headers)
            for row in data_rows:
                padded = list(row) + [None] * max(0, n - len(row))
                records.append([to_str(padded[i]) for i in range(n)])

            df_raw = pd.DataFrame(records, columns=headers)

            # Remove linhas 100% vazias
            df_raw = df_raw.dropna(how="all")
            if df_raw.empty or len(df_raw.columns) < 3:
                continue

            df = normalizar_df(df_raw, ano, mes)
            if df.empty:
                continue

            frames.append(df)
            mes_str = f"mês {mes}" if mes else "mês ?"
            print(f"      📅 {sheet_name:<14} → {len(df):>7,} registros  ({mes_str})")

        except Exception as e:
            print(f"      ⚠️  Aba '{sheet_name}' erro: {type(e).__name__}: {e}")
            print(traceback.format_exc())

    wb.close()
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()


# ──────────────────────────────────────────────
# BANCO — SALVAR
# ──────────────────────────────────────────────

def salvar_no_banco(df: pd.DataFrame, ano: int, hash_arquivo: str):
    PH = _db.PH
    with _db.get_db() as conn:
        _db.execute(conn, f"DELETE FROM transacoes WHERE ano_referencia = {PH}", [ano])
        # Insere linha a linha para compatibilidade PG/SQLite
        cols = [c for c in df.columns]
        placeholders = ", ".join([PH] * len(cols))
        col_names = ", ".join(cols)
        sql = f"INSERT INTO transacoes ({col_names}) VALUES ({placeholders})"
        if _db.USE_PG:
            cur = conn.cursor()
            import psycopg2.extras
            rows = [tuple(None if (isinstance(v, float) and __import__('math').isnan(v)) else v
                         for v in row) for row in df.itertuples(index=False)]
            psycopg2.extras.execute_batch(cur, sql.replace("?", "%s"), rows, page_size=500)
        else:
            import sqlite3 as _sq3
            conn2 = _sq3.connect(str(_db.DB_PATH))
            conn2.execute(f"DELETE FROM transacoes WHERE ano_referencia = ?", (ano,))
            df.to_sql("transacoes", conn2, if_exists="append", index=False)
            conn2.execute("INSERT OR REPLACE INTO arquivos_processados (ano, hash, linhas, updated_at) VALUES (?,?,?,?)",
                         (ano, hash_arquivo, len(df), datetime.now().isoformat()))
            conn2.commit()
            conn2.close()
            return
        _db.execute(conn, f"""
            INSERT INTO arquivos_processados (ano, hash, linhas, updated_at)
            VALUES ({PH},{PH},{PH},{PH})
            ON CONFLICT (ano) DO UPDATE SET hash={PH}, linhas={PH}, updated_at={PH}
        """, [ano, hash_arquivo, len(df), datetime.now().isoformat(),
              hash_arquivo, len(df), datetime.now().isoformat()])


def hash_em_banco(ano: int) -> str | None:
    try:
        with _db.get_db() as conn:
            row = _db.fetchone(conn,
                f"SELECT hash FROM arquivos_processados WHERE ano = {_db.PH}", [ano])
            return row["hash"] if row else None
    except Exception:
        return None


# ──────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────

def sincronizar(anos: list[int] = None, forcar: bool = False):
    init_db()
    anos_alvo = sorted(anos or list(ARQUIVOS.keys()), reverse=True)
    total_geral = 0

    for ano in anos_alvo:
        url = ARQUIVOS.get(ano)
        if not url:
            print(f"⚠️  Ano {ano} sem URL configurada.")
            continue

        cache_path = download(ano, url)
        if cache_path is None:
            continue

        novo_hash = file_hash(cache_path)
        hash_salvo = hash_em_banco(ano)

        if not forcar and novo_hash == hash_salvo:
            print(f"   ↩️  [{ano}] Arquivo não mudou, pulando.\n")
            continue

        print(f"   📊 [{ano}] Processando abas...")
        df = processar_xlsx(cache_path, ano)

        if df.empty:
            print(f"   ❌ [{ano}] Nenhum dado extraído.\n")
            continue

        salvar_no_banco(df, ano, novo_hash)
        total_geral += len(df)
        print(f"   ✅ [{ano}] {len(df):,} registros salvos.\n")

    print(f"\n🎉 Pronto! Total inserido: {total_geral:,} registros.")


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--anos", nargs="+", type=int)
    parser.add_argument("--forcar", action="store_true")
    args = parser.parse_args()
    sincronizar(anos=args.anos, forcar=args.forcar)
