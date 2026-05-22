"""
exportar.py — Geração de Excel e PDF a partir dos dados ITBI-SP
"""
import io
from pathlib import Path
from datetime import datetime
from typing import Optional

import pandas as pd

import db as _db

# Colunas na ordem da planilha original do governo
COLUNAS = [
    ("sql_terreno",           "N° do Cadastro (SQL)"),
    ("logradouro",            "Nome do Logradouro"),
    ("numero",                "Número"),
    ("complemento",           "Complemento"),
    ("bairro",                "Bairro"),
    ("cep",                   "CEP"),
    ("natureza_transacao",    "Natureza de Transação"),
    ("valor_declarado",       "Valor de Transação (declarado pelo contribuinte)"),
    ("data_transacao",        "Data de Transação"),
    ("valor_venal_ref",       "Valor Venal de Referência"),
    ("proporcao_transmitida", "Proporção Transmitida (%)"),
    ("valor_itbi",            "Base de Cálculo adotada"),
    ("tipo_financiamento",    "Tipo de Financiamento"),
    ("valor_financiado",      "Valor Financiado"),
    ("cartorio_registro",     "Cartório de Registro"),
    ("matricula_imovel",      "Matrícula do Imóvel"),
    ("situacao_sql",          "Situação do SQL"),
    ("area_terreno",          "Área do Terreno (m2)"),
    ("testada",               "Testada (m)"),
    ("fracao_ideal",          "Fração Ideal"),
    ("area_construida",       "Área Construída (m2)"),
    ("tipo_uso",              "Uso (IPTU)"),
    ("descricao_uso",         "Descrição do uso (IPTU)"),
    ("padrao_iptu",           "Padrão (IPTU)"),
    ("acc_iptu",              "ACC (IPTU)"),
    ("ano_referencia",        "Ano Referência"),
    ("mes_referencia",        "Mês Referência"),
]

COLS_MONETARIAS = {"valor_declarado", "valor_financiado", "valor_itbi", "valor_venal_ref"}
COLS_NUMERICAS  = {"area_terreno", "area_construida", "testada", "proporcao_transmitida"}


# ──────────────────────────────────────────────
# CONSULTA
# ──────────────────────────────────────────────

def buscar(filtros: dict, ids: list = None, limit: int = 100_000) -> pd.DataFrame:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row

    if ids:
        ph = ','.join('?' * len(ids))
        rows = conn.execute(
            f"SELECT * FROM transacoes WHERE id IN ({ph}) ORDER BY ano_referencia DESC, mes_referencia DESC",
            ids
        ).fetchall()
    else:
        conds, params = [], []
        if filtros.get('logradouro'):
            conds.append("UPPER(logradouro) LIKE UPPER(?)")
            params.append(f"%{filtros['logradouro']}%")
        if filtros.get('numero'):
            conds.append("numero = ?")
            params.append(filtros['numero'].strip())
        if filtros.get('bairro'):
            conds.append("UPPER(bairro) LIKE UPPER(?)")
            params.append(f"%{filtros['bairro']}%")
        if filtros.get('cep'):
            conds.append("cep LIKE ?")
            params.append(f"{filtros['cep'].replace('-','')}%")
        if filtros.get('ano_min'):
            conds.append("ano_referencia >= ?")
            params.append(int(filtros['ano_min']))
        if filtros.get('ano_max'):
            conds.append("ano_referencia <= ?")
            params.append(int(filtros['ano_max']))
        if filtros.get('valor_min'):
            conds.append("valor_declarado >= ?")
            params.append(float(filtros['valor_min']))
        if filtros.get('valor_max'):
            conds.append("valor_declarado <= ?")
            params.append(float(filtros['valor_max']))
        if filtros.get('natureza'):
            conds.append("UPPER(natureza_transacao) LIKE UPPER(?)")
            params.append(f"%{filtros['natureza']}%")

        where = ("WHERE " + " AND ".join(conds)) if conds else ""
        rows = conn.execute(
            f"SELECT * FROM transacoes {where} ORDER BY ano_referencia DESC, mes_referencia DESC LIMIT {limit}",
            params
        ).fetchall()

    conn.close()
    return pd.DataFrame([dict(r) for r in rows]) if rows else pd.DataFrame()


# ──────────────────────────────────────────────
# EXCEL
# ──────────────────────────────────────────────

def gerar_excel(df: pd.DataFrame, filtros: dict = None) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1

    wb = Workbook()

    # ── Aba 1: Dados ──────────────────────────
    ws = wb.active
    ws.title = "ITBI - Dados"

    AZUL_ESCURO = "1A2F6B"
    AZUL_CLARO  = "E8EDFF"
    CINZA_LINHA = "F5F7FF"

    side = Side(style='thin', color='C5CCE8')
    borda = Border(left=side, right=side, top=side, bottom=side)

    # Cabeçalho
    for ci, (_, nome) in enumerate(COLUNAS, 1):
        cell = ws.cell(row=1, column=ci, value=nome)
        cell.font      = Font(bold=True, color="FFFFFF", size=9, name="Calibri")
        cell.fill      = PatternFill("solid", fgColor=AZUL_ESCURO)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = borda
    ws.row_dimensions[1].height = 40

    # Dados
    col_keys = [c for c, _ in COLUNAS]
    for ri, row in enumerate(df.itertuples(index=False), 2):
        bg = CINZA_LINHA if ri % 2 == 0 else "FFFFFF"
        for ci, ck in enumerate(col_keys, 1):
            val = getattr(row, ck, None) if ck in df.columns else None
            # Converte NaN/None
            if pd.isna(val) if isinstance(val, float) else val is None:
                val = None
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.border    = borda
            cell.alignment = Alignment(vertical="center")
            if ck in COLS_MONETARIAS and val is not None:
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif ck in COLS_NUMERICAS and val is not None:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")

    # Larguras
    larguras = {
        "sql_terreno": 20, "logradouro": 32, "numero": 8, "complemento": 16,
        "bairro": 22, "cep": 12, "natureza_transacao": 35,
        "valor_declarado": 22, "data_transacao": 15,
        "valor_venal_ref": 22, "proporcao_transmitida": 16,
        "valor_itbi": 22, "tipo_financiamento": 22, "valor_financiado": 22,
        "cartorio_registro": 25, "matricula_imovel": 18,
        "situacao_sql": 15, "area_terreno": 16, "testada": 12,
        "fracao_ideal": 14, "area_construida": 16,
        "tipo_uso": 12, "descricao_uso": 25, "padrao_iptu": 14,
        "acc_iptu": 12, "ano_referencia": 10, "mes_referencia": 8,
    }
    for ci, (ck, _) in enumerate(COLUNAS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = larguras.get(ck, 14)

    ws.freeze_panes = "A2"

    # ── Aba 2: Resumo ─────────────────────────
    ws2 = wb.create_sheet("Resumo")

    def vbr(v):
        if v is None or (isinstance(v, float) and pd.isna(v)): return "—"
        return f"R$ {int(round(v)):,}".replace(",", ".")

    def estilo_resumo(cell, negrito=False, cor_fundo=None, cor_texto="000000"):
        cell.font = Font(bold=negrito, size=10, name="Calibri", color=cor_texto)
        if cor_fundo:
            cell.fill = PatternFill("solid", fgColor=cor_fundo)
        cell.border = borda
        cell.alignment = Alignment(vertical="center", indent=1)

    titulo = ws2.cell(1, 1, "Relatório ITBI · São Paulo")
    titulo.font = Font(bold=True, size=16, color="1A2F6B", name="Calibri")
    ws2.merge_cells("A1:C1")
    ws2.row_dimensions[1].height = 30

    ws2.cell(2, 1, f"Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}")
    ws2.merge_cells("A2:C2")

    linhas_resumo = [
        ("", "", ""),
        ("MÉTRICAS", "Valor", ""),
        ("Total de transações", f"{len(df):,}".replace(",", "."), ""),
        ("Volume total (R$)", vbr(df["valor_declarado"].sum() if "valor_declarado" in df.columns else 0), ""),
        ("Ticket médio (R$)", vbr(df["valor_declarado"].mean() if "valor_declarado" in df.columns else 0), ""),
        ("Maior transação (R$)", vbr(df["valor_declarado"].max() if "valor_declarado" in df.columns else 0), ""),
        ("Menor transação (R$)", vbr(df["valor_declarado"].min() if "valor_declarado" in df.columns else 0), ""),
        ("Base ITBI total (R$)", vbr(df["valor_itbi"].sum() if "valor_itbi" in df.columns else 0), ""),
    ]
    if filtros:
        linhas_resumo.append(("", "", ""))
        linhas_resumo.append(("FILTROS APLICADOS", "", ""))
        for k, v in filtros.items():
            if v: linhas_resumo.append((k, str(v), ""))

    for ri, (a, b, c) in enumerate(linhas_resumo, 4):
        ca = ws2.cell(ri, 1, a)
        cb = ws2.cell(ri, 2, b)
        negrito = a in ("MÉTRICAS", "FILTROS APLICADOS")
        bg = AZUL_ESCURO if negrito else ("FFFFFF" if ri % 2 == 0 else CINZA_LINHA)
        cor_tx = "FFFFFF" if negrito else "000000"
        estilo_resumo(ca, negrito, bg, cor_tx)
        estilo_resumo(cb, negrito, bg, cor_tx)

    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 28

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ──────────────────────────────────────────────
# PDF
# ──────────────────────────────────────────────

def gerar_pdf(df: pd.DataFrame, filtros: dict = None) -> bytes:
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import matplotlib.ticker as mticker
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                     Table, TableStyle, Image as RLImage,
                                     HRFlowable, KeepTogether)
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm,
        title="Relatório ITBI · São Paulo"
    )

    C_AZUL    = colors.HexColor('#1A2F6B')
    C_ACCENT  = colors.HexColor('#5b7fff')
    C_MUTED   = colors.HexColor('#6b748f')
    C_BG      = colors.HexColor('#f0f3fa')
    C_LINHA1  = colors.white
    C_LINHA2  = colors.HexColor('#f5f7ff')
    C_BORDA   = colors.HexColor('#dde3f5')

    styles = getSampleStyleSheet()
    titulo_s   = ParagraphStyle('T', fontSize=22, fontName='Helvetica-Bold',
                                 textColor=C_AZUL, spaceAfter=4)
    subtitulo_s = ParagraphStyle('S', fontSize=9,  fontName='Helvetica',
                                  textColor=C_MUTED, spaceAfter=4)
    secao_s    = ParagraphStyle('SE', fontSize=11, fontName='Helvetica-Bold',
                                 textColor=C_ACCENT, spaceBefore=14, spaceAfter=6)
    nota_s     = ParagraphStyle('N', fontSize=7, fontName='Helvetica',
                                 textColor=C_MUTED, spaceAfter=4)

    elems = []
    W = 27.7 * cm  # largura útil A4 paisagem

    # ── Cabeçalho ──────────────────────────────
    elems.append(Paragraph("Dashboard ITBI · São Paulo", titulo_s))
    filtros_str = "  |  ".join(f"<b>{k}:</b> {v}" for k, v in (filtros or {}).items() if v)
    sub = f"Gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')}"
    if filtros_str:
        sub += f"  ·  Filtros aplicados: {filtros_str}"
    elems.append(Paragraph(sub, subtitulo_s))
    elems.append(HRFlowable(width=W, thickness=1.5, color=C_AZUL, spaceAfter=10))

    # ── KPIs ────────────────────────────────────
    elems.append(Paragraph("Resumo Estatístico", secao_s))

    total = len(df)
    def vbr(v):
        if v is None or (isinstance(v, float) and pd.isna(v)) or v == 0: return "—"
        return f"R$ {int(round(v)):,}".replace(",", ".")

    vol  = df['valor_declarado'].sum()  if 'valor_declarado' in df.columns else 0
    med  = df['valor_declarado'].mean() if 'valor_declarado' in df.columns else 0
    maxi = df['valor_declarado'].max()  if 'valor_declarado' in df.columns else 0
    mini = df['valor_declarado'].min()  if 'valor_declarado' in df.columns else 0
    itbi = df['valor_itbi'].sum()       if 'valor_itbi'      in df.columns else 0

    kpi_labels = ["Total de Transações", "Volume Total", "Ticket Médio",
                  "Maior Transação", "Menor Transação", "Base ITBI Total"]
    kpi_vals   = [f"{total:,}".replace(',','.'), vbr(vol), vbr(med),
                  vbr(maxi), vbr(mini), vbr(itbi)]

    kpi_table = Table(
        [kpi_labels, kpi_vals],
        colWidths=[W/6]*6
    )
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND',   (0,0), (-1,0), C_AZUL),
        ('TEXTCOLOR',    (0,0), (-1,0), colors.white),
        ('FONTNAME',     (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',     (0,0), (-1,0), 8),
        ('ALIGN',        (0,0), (-1,-1), 'CENTER'),
        ('BACKGROUND',   (0,1), (-1,1), C_BG),
        ('FONTNAME',     (0,1), (-1,1), 'Helvetica-Bold'),
        ('FONTSIZE',     (0,1), (-1,1), 11),
        ('TEXTCOLOR',    (0,1), (-1,1), C_ACCENT),
        ('GRID',         (0,0), (-1,-1), 0.5, C_BORDA),
        ('TOPPADDING',   (0,0), (-1,-1), 8),
        ('BOTTOMPADDING',(0,0), (-1,-1), 8),
        ('ROWBACKGROUNDS',(0,1),(-1,1), [C_BG]),
    ]))
    elems.append(kpi_table)
    elems.append(Spacer(1, 14))

    # ── Gráficos ────────────────────────────────
    if len(df) > 0 and 'ano_referencia' in df.columns:
        elems.append(Paragraph("Análise Visual", secao_s))

        fig, axes = plt.subplots(1, 2, figsize=(15, 4.2))
        fig.patch.set_facecolor('#f8f9ff')

        paleta = ['#5b7fff','#a78bfa','#2dd4a0','#f5c842','#f26c6c',
                  '#22d3ee','#f472b6','#fb923c','#a3e635','#e879f9']

        # Gráfico 1 — por ano
        ax1 = axes[0]
        por_ano = df.groupby('ano_referencia').size().reset_index(name='n').sort_values('ano_referencia')
        ax1.bar(por_ano['ano_referencia'].astype(str), por_ano['n'],
                color='#5b7fff', alpha=0.85, edgecolor='#3b5fd0', linewidth=0.4, zorder=3)
        ax1.set_title('Transações por Ano', fontsize=11, fontweight='bold',
                      color='#1A2F6B', pad=8)
        ax1.tick_params(axis='x', rotation=45, labelsize=8)
        ax1.tick_params(axis='y', labelsize=8)
        ax1.set_facecolor('#f0f3fa')
        ax1.grid(axis='y', alpha=0.5, color='#dde3f5', zorder=0)
        ax1.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f'{int(x):,}'.replace(',','.')))
        for sp in ['top','right']: ax1.spines[sp].set_visible(False)

        # Gráfico 2 — top bairros
        ax2 = axes[1]
        if 'bairro' in df.columns:
            top_b = df[df['bairro'].notna()].groupby('bairro').size().nlargest(10).reset_index(name='n')
            ax2.barh(top_b['bairro'][::-1], top_b['n'][::-1],
                     color=paleta[:len(top_b)], alpha=0.9, edgecolor='none', zorder=3)
            ax2.set_title('Top 10 Bairros', fontsize=11, fontweight='bold',
                          color='#1A2F6B', pad=8)
            ax2.tick_params(axis='both', labelsize=7)
            ax2.set_facecolor('#f0f3fa')
            ax2.grid(axis='x', alpha=0.5, color='#dde3f5', zorder=0)
            for sp in ['top','right']: ax2.spines[sp].set_visible(False)

        plt.tight_layout(pad=2.5)
        imgbuf = io.BytesIO()
        plt.savefig(imgbuf, format='png', dpi=140, bbox_inches='tight',
                    facecolor='#f8f9ff', edgecolor='none')
        plt.close(fig)
        imgbuf.seek(0)
        elems.append(RLImage(imgbuf, width=W, height=8.5*cm))
        elems.append(Spacer(1, 14))

    # ── Tabela de dados ─────────────────────────
    MAX_ROWS = 500
    elems.append(Paragraph(
        f"Registros ({total:,} transações{f' · exibindo os primeiros {MAX_ROWS}' if total > MAX_ROWS else ''})".replace(',','.'),
        secao_s
    ))

    HD = ["Ano/Mês", "Data", "Logradouro", "Nº", "Bairro",
          "Valor Declarado (R$)", "Base ITBI (R$)", "Área Constr.", "Natureza"]
    CW = [2.1*cm, 2.4*cm, 5.8*cm, 1.2*cm, 4.2*cm, 3.8*cm, 3.4*cm, 2.2*cm, 4.6*cm]

    def trunc(v, n):
        s = str(v or '').strip()
        return s[:n-1]+'…' if len(s) > n else s

    rows_pdf = [HD]
    for _, r in df.head(MAX_ROWS).iterrows():
        am = str(int(r.get('ano_referencia') or 0))
        if r.get('mes_referencia'):
            am += f"/{int(r['mes_referencia']):02d}"
        nat = trunc(str(r.get('natureza_transacao') or '').split('.',1)[-1].strip() or r.get('natureza_transacao',''), 30)
        vd = vbr(r.get('valor_declarado'))
        vi = vbr(r.get('valor_itbi'))
        ac = f"{r.get('area_construida','') or ''}"
        rows_pdf.append([am, trunc(r.get('data_transacao',''),14), trunc(r.get('logradouro',''),30),
                         trunc(r.get('numero',''),6), trunc(r.get('bairro',''),22),
                         vd, vi, ac, nat])

    if total > MAX_ROWS:
        rows_pdf.append([Paragraph(f"<i>... e mais {total-MAX_ROWS:,} registros não exibidos</i>".replace(',','.'), nota_s)] + ['']*8)

    t = Table(rows_pdf, colWidths=CW, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND',    (0,0), (-1,0), C_AZUL),
        ('TEXTCOLOR',     (0,0), (-1,0), colors.white),
        ('FONTNAME',      (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',      (0,0), (-1,0), 7.5),
        ('ALIGN',         (0,0), (-1,0), 'CENTER'),
        ('ROWBACKGROUNDS',(0,1), (-1,-1), [C_LINHA1, C_LINHA2]),
        ('FONTSIZE',      (0,1), (-1,-1), 7),
        ('FONTNAME',      (0,1), (-1,-1), 'Helvetica'),
        ('GRID',          (0,0), (-1,-1), 0.3, C_BORDA),
        ('ALIGN',         (5,1), (6,-1), 'RIGHT'),
        ('FONTNAME',      (5,1), (6,-1), 'Helvetica-Bold'),
        ('TOPPADDING',    (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('LEFTPADDING',   (0,0), (-1,-1), 4),
        ('RIGHTPADDING',  (0,0), (-1,-1), 4),
        ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
    ]))
    elems.append(t)

    # Rodapé de nota
    elems.append(Spacer(1, 8))
    elems.append(Paragraph(
        "Fonte: Secretaria Municipal de Finanças de São Paulo · prefeitura.sp.gov.br/web/fazenda · "
        "Dashboard ITBI-SP",
        nota_s
    ))

    doc.build(elems)
    buf.seek(0)
    return buf.getvalue()
