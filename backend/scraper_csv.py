"""
scraper_csv.py — Versão otimizada do scraper ITBI para VMs com pouca RAM.

Diferenças em relação ao scraper.py original:
  - Default cobre apenas 2024–2026 (carga inicial leve para VMs com pouca RAM)
  - Anos de 2006–2023 ficam disponíveis sob demanda via --anos
  - Lê cada aba do XLSX com pandas (sem openpyxl linha a linha)
  - Converte cada aba para CSV e faz cache em backend/cache/csv/
  - Reruns leem direto do CSV → muito mais rápido e menos RAM
  - Processa e insere uma aba por vez para liberar memória entre iterações

Uso:
  python scraper_csv.py               # baixa/atualiza 2024–2026 (default)
  python scraper_csv.py --anos 2023 2024 2025 2026   # estende para 2023
  python scraper_csv.py --anos 2006 2007 2008 ... 2026  # carga histórica completa
  python scraper_csv.py --forcar      # re-converte mesmo sem mudança de hash
  python scraper_csv.py --limpar-csv  # apaga CSVs em cache antes de processar
"""

import gc
import warnings
warnings.filterwarnings("ignore", category=FutureWarning)
import re
import hashlib
import sqlite3
import argparse
import requests
import pandas as pd
from datetime import datetime
from pathlib import Path

# ──────────────────────────────────────────────
DB_PATH   = Path(__file__).parent / "itbi.db"
CACHE_DIR = Path(__file__).parent / "cache"
CSV_DIR   = Path(__file__).parent / "cache" / "csv"
CACHE_DIR.mkdir(exist_ok=True)
CSV_DIR.mkdir(exist_ok=True)

# Default de sincronizacao = 2024-2026 (carga inicial em VMs com pouca RAM).
# Anos anteriores ficam disponiveis sob demanda via --anos (ex.: --anos 2023 2022 ... 2006).
ANOS_DEFAULT = [2026, 2025, 2024]

ARQUIVOS = {
    2026: "https://prefeitura.sp.gov.br/documents/d/fazenda/guias-de-itbi-pagas-4-xlsx",
    2025: "https://prefeitura.sp.gov.br/cidade/secretarias/upload/fazenda/arquivos/itbi/GUIAS%20DE%20ITBI%20PAGAS%20%2828012026%29%20XLS.xlsx",
    2024: "https://prefeitura.sp.gov.br/cidade/secretarias/upload/fazenda/arquivos/itbi/GUIAS-DE-ITBI-PAGAS-2024.xlsx",
    2023: "https://www.prefeitura.sp.gov.br/cidade/secretarias/upload/fazenda/arquivos/XLSX/GUIAS-DE-ITBI-PAGAS-2023.xlsx",
    2022: "https://www.prefeitura.sp.gov.br/cidade/secretarias/upload/fazenda/arquivos/XLSX/GUIAS_DE_ITBI_PAGAS_12-2022.xlsx",
    2021: "https://www.prefeitura.sp.gov.br/cidade/secretarias/upload/fazenda/arquivos/itbi/ITBI_Setembro_2022/GUIAS_DE_ITBI_PAGAS_(2021).xlsx",
    2020: "https://www.prefeitura.sp.gov.br/cidade/secretarias/upload/fazenda/arquivos/itbi/ITBI_Setembro_2022/GUIAS_DE_ITBI_PAGAS_(2020).xlsx",
    2019: "https://www.prefeitura.sp.gov.br/cidade/secretarias/upload/fazenda/arquivos/itbi/GUIAS_DE_ITBI_PAGAS_(2019).xlsx",
    2018: "https://www.prefeitura.sp.gov.br/cidade/secretarias/upload/fazenda/arquivos/itbi/guias_de_itbi_pagas_2018.xlsx",
    2017: "https://www.prefeitura.sp.gov.br/cidade/secretarias/upload/fazenda/arquivos/itbi/guias_de_itbi_pagas_2017.xlsx",
    2016: "https://www.prefeitura.sp.gov.br/cidade/secretarias/upload/fazenda/arquivos/itbi/guias_de_itbi_pagas_2016.xlsx",
    2015: "https://www.prefeitura.sp.gov.br/cidade/secretarias/upload/fazenda/arquivos/itbi/guias_de_itbi_pagas_2015.xlsx",
    2014: "https://www.prefeitura.sp.gov.br/cidade/secretarias/upload/fazenda/arquivos/itbi/guias_de_itbi_pagas_2014.xlsx",
    2013: "https://www.prefeitura.sp.gov.br/cidade/secretarias/upload/fazenda/arquivos/itbi/guias_de_itbi_pagas_2013.xlsx",
    2012: "https://www.prefeitura.sp.gov.br/cidade/secretarias/upload/fazenda/arquivos/itbi/guias_de_itbi_pagas_2012.xlsx",
    2011: "https://www.prefeitura.sp.gov.br/cidade/secretarias/upload/fazenda/arquivos/itbi/guias_de_itbi_pagas_2011.xlsx",
    2010: "https://www.prefeitura.sp.gov.br/cidade/secretarias/upload/fazenda/arquivos/itbi/guias_de_itbi_pagas_2010.xlsx",
    2009: "https://www.prefeitura.sp.gov.br/cidade/secretarias/upload/fazenda/arquivos/itbi/guias_de_itbi_pagas_2009.xlsx",
    2008: "https://www.prefeitura.sp.gov.br/cidade/secretarias/upload/fazenda/arquivos/itbi/guias_de_itbi_pagas_2008.xlsx",
    2007: "https://www.prefeitura.sp.gov.br/cidade/secretarias/upload/fazenda/arquivos/itbi/guias_de_itbi_pagas_2007.xlsx",
    2006: "https://www.prefeitura.sp.gov.br/cidade/secretarias/upload/fazenda/arquivos/itbi/guias_de_itbi_pagas_2006.xlsx",
}

HEADERS = {"User-Agent": "Mozilla/5.0 (compatible; ITBI-Dashboard/1.0)"}

PAGE_ITBI = "https://prefeitura.sp.gov.br/web/fazenda/w/acesso_a_informacao/31501"


def buscar_url_dinamica(ano: int) -> str | None:
    """Raspa a página oficial da Prefeitura SP e retorna a URL mais recente do XLSX do ano."""
    import re as _re
    from urllib.parse import urljoin, quote
    try:
        hdrs = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                          "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "pt-BR,pt;q=0.9",
            "Referer": "https://prefeitura.sp.gov.br/",
        }
        r = requests.get(PAGE_ITBI, headers=hdrs, timeout=30)
        r.raise_for_status()
        html = r.text

        # 1. Links /documents/d/fazenda/ com xlsx no nome (padrão Liferay mais recente)
        padrao_doc = _re.compile(
            r'href=["\'](/documents/d/fazenda/[^"\']*(?:itbi|guias)[^"\']*xlsx?[^"\']*)["\']',
            _re.IGNORECASE
        )
        matches = padrao_doc.findall(html)

        if not matches:
            # 2. Links .xlsx/.xls que contenham o ano
            padrao = _re.compile(
                r'href=["\']([^"\']*' + str(ano) + r'[^"\']*\.xlsx?)["\']',
                _re.IGNORECASE
            )
            matches = padrao.findall(html)

        if not matches:
            # 3. Fallback: qualquer link com itbi + ano
            padrao2 = _re.compile(
                r'href=["\']([^"\']*(?:itbi|ITBI)[^"\']*' + str(ano) + r'[^"\']*)["\']',
                _re.IGNORECASE
            )
            matches = padrao2.findall(html)

        if not matches:
            print(f"⚠️  [{ano}] Nenhum link encontrado na página para {ano}.")
            return None

        # Prefere o link com data mais recente: extrai números no formato DDMMYYYY do nome
        def _data_no_nome(u):
            nums = _re.findall(r'\d{8}', u)
            for n in nums:
                try:
                    from datetime import datetime as _dt2
                    return _dt2.strptime(n, "%d%m%Y")
                except Exception:
                    pass
            return None

        urls_com_data = [(u, _data_no_nome(u)) for u in matches]
        urls_com_data.sort(key=lambda x: x[1] or __import__('datetime').datetime.min, reverse=True)
        url = urls_com_data[0][0]

        # Garante URL absoluta e espaços encodados
        if url.startswith('/'):
            url = 'https://prefeitura.sp.gov.br' + url
        # Encoda espaços (mas preserva % já encodados)
        url = _re.sub(r'(?<!%)(?<!\%[0-9A-Fa-f]) ', '%20', url)
        url = url.replace(' ', '%20')

        print(f"🔗 [{ano}] URL mais recente: {url}")
        return url

    except Exception as e:
        print(f"⚠️  [{ano}] Falha ao raspar página: {e}")
        return None

MESES_PT = {
    "jan": 1, "fev": 2, "mar": 3, "abr": 4,
    "mai": 5, "jun": 6, "jul": 7, "ago": 8,
    "set": 9, "out": 10, "nov": 11, "dez": 12,
}

IGNORAR = {"legenda", "sumario", "sumário", "instrucao", "leia",
           "orientacao", "explicac", "explicacao", "explicações", "explicacoes",
           "tabela de uso", "tabela de usos", "tabela de padrao", "tabela de padroes",
           "tabela de padrões", "usos", "padroes", "padrões", "uso iptu", "padrao iptu"}


# ──────────────────────────────────────────────
# BANCO DE DADOS (igual ao scraper.py original)
# ──────────────────────────────────────────────

def init_db():
    conn = sqlite3.connect(DB_PATH)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS transacoes (
            id                    INTEGER PRIMARY KEY AUTOINCREMENT,
            ano_referencia        INTEGER,
            mes_referencia        INTEGER,
            data_transacao        TEXT,
            logradouro            TEXT,
            numero                TEXT,
            complemento           TEXT,
            bairro                TEXT,
            cep                   TEXT,
            sql_terreno           TEXT,
            area_terreno          REAL,
            area_construida       REAL,
            valor_declarado       REAL,
            valor_financiado      REAL,
            valor_itbi            REAL,
            natureza_transacao    TEXT,
            tipo_uso              TEXT,
            descricao_uso         TEXT,
            proporcao_transmitida REAL,
            tipo_financiamento    TEXT,
            acc_iptu              TEXT,
            valor_venal_ref       REAL,
            cartorio_registro     TEXT,
            matricula_imovel      TEXT,
            situacao_sql          TEXT,
            testada               REAL,
            fracao_ideal          TEXT,
            padrao_iptu           TEXT,
            created_at            TEXT DEFAULT (datetime('now'))
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS arquivos_processados (
            ano        INTEGER PRIMARY KEY,
            hash       TEXT,
            linhas     INTEGER,
            updated_at TEXT
        )
    """)
    novas_colunas = [
        ("descricao_uso",         "TEXT"),
        ("proporcao_transmitida",  "REAL"),
        ("tipo_financiamento",     "TEXT"),
        ("acc_iptu",               "TEXT"),
        ("valor_venal_ref",        "REAL"),
        ("cartorio_registro",      "TEXT"),
        ("matricula_imovel",       "TEXT"),
        ("situacao_sql",           "TEXT"),
        ("testada",                "REAL"),
        ("fracao_ideal",           "TEXT"),
        ("padrao_iptu",            "TEXT"),
    ]
    cols_existentes = {row[1] for row in conn.execute("PRAGMA table_info(transacoes)")}
    for nome, tipo in novas_colunas:
        if nome not in cols_existentes:
            conn.execute(f"ALTER TABLE transacoes ADD COLUMN {nome} {tipo}")
    conn.commit()
    conn.close()
    print("✅ Banco pronto:", DB_PATH)


# ──────────────────────────────────────────────
# DOWNLOAD
# ──────────────────────────────────────────────

def file_hash(path: Path) -> str:
    h = hashlib.md5()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


def download(ano: int, url: str) -> Path | None:
    cache_path = CACHE_DIR / f"itbi_{ano}.xlsx"
    # Garante que espaços na URL estejam encodados
    url = url.replace(' ', '%20')
    print(f"📥 [{ano}] Baixando {url[:80]}...", end=" ", flush=True)
    hdrs = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                      "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
        "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,"
                  "application/vnd.ms-excel,*/*",
        "Accept-Language": "pt-BR,pt;q=0.9",
        "Referer": PAGE_ITBI,
    }
    try:
        sess = requests.Session()
        # Visita a página pai primeiro para obter cookies de sessão
        sess.get(PAGE_ITBI, headers=hdrs, timeout=15)
        r = sess.get(url, headers=hdrs, timeout=300, stream=True, allow_redirects=True)
        r.raise_for_status()
        # Verifica se recebeu arquivo binário (não HTML de erro)
        ct = r.headers.get('Content-Type', '')
        if 'html' in ct and r.headers.get('Content-Length', '99999') == '0':
            raise ValueError(f"Servidor retornou HTML em vez do arquivo. Content-Type: {ct}")
        with open(cache_path, "wb") as f:
            for chunk in r.iter_content(chunk_size=65536):
                f.write(chunk)
        size_kb = cache_path.stat().st_size / 1024
        if size_kb < 10:
            cache_path.unlink(missing_ok=True)
            raise ValueError(f"Arquivo muito pequeno ({size_kb:.0f} KB) — provável página de erro")
        print(f"✅ ({size_kb:.0f} KB)")
        return cache_path
    except Exception as e:
        print(f"❌ {e}")
        return None


# ──────────────────────────────────────────────
# NORMALIZAÇÃO (igual ao scraper.py original)
# ──────────────────────────────────────────────

def limpar_col(col: str) -> str:
    col = str(col).lower().strip()
    col = col.translate(str.maketrans(
        'áàãâäéèêëíìîïóòõôöúùûüçñ',
        'aaaaaeeeeiiiiooooouuuucn'
    ))
    col = re.sub(r'[^\w]', '_', col)
    col = re.sub(r'_+', '_', col)
    return col.strip('_')


COL_MAP = {
    "n_do_cadastro_sql_":                               "sql_terreno",
    "n_do_cadastro__sql_":                              "sql_terreno",
    "n_do_cadastro_sql":                                "sql_terreno",
    "n_cadastro_sql":                                   "sql_terreno",
    "n_cadastro__sql_":                                 "sql_terreno",
    "ncadastro_sql":                                    "sql_terreno",
    "n_sql":                                            "sql_terreno",
    "sql":                                              "sql_terreno",
    "n_do_cadastro":                                    "sql_terreno",
    "cadastro_sql":                                     "sql_terreno",
    "nro_contribuinte":                                 "sql_terreno",
    "numero_contribuinte":                              "sql_terreno",
    "contribuinte":                                     "sql_terreno",
    "nome_do_logradouro":                               "logradouro",
    "logradouro":                                       "logradouro",
    "nome_logradouro":                                  "logradouro",
    "endereco":                                         "logradouro",
    "numero":                                           "numero",
    "num":                                              "numero",
    "complemento":                                      "complemento",
    "bairro":                                           "bairro",
    "cep":                                              "cep",
    "natureza_de_transacao":                            "natureza_transacao",
    "natureza_transacao":                               "natureza_transacao",
    "natureza_da_transacao":                            "natureza_transacao",
    "natureza":                                         "natureza_transacao",
    "valor_de_transacao_declarado_pelo_contribuinte_":  "valor_declarado",
    "valor_de_transacao_declarado_pelo_contribuinte":   "valor_declarado",
    "valor_de_transacao__declarado_pelo_contribuinte_": "valor_declarado",
    "valor_declarado":                                  "valor_declarado",
    "valor_transacao":                                  "valor_declarado",
    "vl_declarado":                                     "valor_declarado",
    "base_de_calculo_adotada":                          "valor_itbi",
    "base_calculo_adotada":                             "valor_itbi",
    "valor_venal_de_referencia_proporcional_":          "valor_itbi",
    "valor_venal_de_referencia_proporcional":           "valor_itbi",
    "valor_itbi":                                       "valor_itbi",
    "vl_itbi":                                          "valor_itbi",
    "proporcao_transmitida____":                        "proporcao_transmitida",
    "proporcao_transmitida___":                         "proporcao_transmitida",
    "proporcao_transmitida__":                          "proporcao_transmitida",
    "proporcao_transmitida_":                           "proporcao_transmitida",
    "proporcao_transmitida":                            "proporcao_transmitida",
    "valor_financiado":                                 "valor_financiado",
    "vl_financiado":                                    "valor_financiado",
    "data_de_transacao":                                "data_transacao",
    "data_transacao":                                   "data_transacao",
    "dt_transacao":                                     "data_transacao",
    "area_do_terreno_m2_":                              "area_terreno",
    "area_do_terreno__m2_":                             "area_terreno",
    "area_terreno":                                     "area_terreno",
    "area_construida_m2_":                              "area_construida",
    "area_construida__m2_":                             "area_construida",
    "area_construida":                                  "area_construida",
    "uso_iptu_":                                        "tipo_uso",
    "uso__iptu_":                                       "tipo_uso",
    "uso_iptu":                                         "tipo_uso",
    "tipo_uso":                                         "tipo_uso",
    "uso":                                              "tipo_uso",
    "descricao_do_uso_iptu_":                           "descricao_uso",
    "descricao_do_uso__iptu_":                          "descricao_uso",
    "descricao_do_uso_iptu":                            "descricao_uso",
    "tipo_de_financiamento":                            "tipo_financiamento",
    "tipo_financiamento":                               "tipo_financiamento",
    "acc_iptu_":                                        "acc_iptu",
    "acc__iptu_":                                       "acc_iptu",
    "acc_iptu":                                         "acc_iptu",
    "acc":                                              "acc_iptu",
    "valor_venal_de_referencia":                        "valor_venal_ref",
    "valor_venal_referencia":                           "valor_venal_ref",
    "cartorio_de_registro":                             "cartorio_registro",
    "cartorio_registro":                                "cartorio_registro",
    "matricula_do_imovel":                              "matricula_imovel",
    "matricula_imovel":                                 "matricula_imovel",
    "situacao_do_sql":                                  "situacao_sql",
    "situacao_sql":                                     "situacao_sql",
    "testada_m_":                                       "testada",
    "testada__m_":                                      "testada",
    "testada":                                          "testada",
    "fracao_ideal":                                     "fracao_ideal",
    "padrao_iptu_":                                     "padrao_iptu",
    "padrao__iptu_":                                    "padrao_iptu",
    "padrao_iptu":                                      "padrao_iptu",
    "descricao_do_padrao_iptu_":                        "padrao_iptu",
}

COLUNAS_FINAIS = [
    "ano_referencia", "mes_referencia", "data_transacao",
    "logradouro", "numero", "complemento", "bairro", "cep",
    "sql_terreno", "area_terreno", "area_construida",
    "valor_declarado", "valor_financiado", "valor_itbi",
    "valor_venal_ref", "proporcao_transmitida",
    "natureza_transacao", "tipo_financiamento",
    "tipo_uso", "descricao_uso", "acc_iptu",
    "cartorio_registro", "matricula_imovel", "situacao_sql",
    "testada", "fracao_ideal", "padrao_iptu",
]

COLS_POSICIONAIS = [
    "N° do Cadastro (SQL)", "Nome do Logradouro", "Número",
    "Complemento", "Bairro", "Referência", "CEP",
    "Natureza de Transação",
    "Valor de Transação (declarado pelo contribuinte)",
    "Data de Transação", "Valor Venal de Referência",
    "Proporção Transmitida (%)",
    "Valor Venal de Referência Proporcional",
    "Tipo de Financiamento", "Valor Financiado",
    "Cartório de Registro", "Matrícula do Imóvel",
    "Situação do SQL", "Área do Terreno (m2)", "Testada (m)",
    "Fração Ideal", "Área Construída (m2)", "Uso (IPTU)",
    "Descrição do Uso (IPTU)", "Padrão (IPTU)", "ACC (IPTU)",
]


def mes_da_aba(nome_aba: str) -> int | None:
    nome = nome_aba.lower().strip()
    for sigla, num in MESES_PT.items():
        if nome.startswith(sigla):
            return num
    return None


def ano_da_aba(nome_aba: str) -> int | None:
    """Extrai o ano (20xx) do nome da aba, ex.: JAN-2025 -> 2025."""
    m = re.search(r'(20\d{2})', str(nome_aba))
    return int(m.group(1)) if m else None


def normalizar_df(df: pd.DataFrame, ano: int, mes: int | None) -> pd.DataFrame:
    df.columns = [limpar_col(c) for c in df.columns]
    df = df.loc[:, ~df.columns.duplicated()]

    if df.columns[0] not in COL_MAP and 'sql_terreno' not in [COL_MAP.get(c) for c in df.columns]:
        COL_MAP[df.columns[0]] = 'sql_terreno'

    for col in df.columns:
        if col not in COL_MAP:
            if col.startswith("valor_de_transacao"):
                COL_MAP[col] = "valor_declarado"
            elif col.startswith("proporcao_transmitida"):
                COL_MAP[col] = "proporcao_transmitida"
            elif col.startswith("area_do_terreno"):
                COL_MAP[col] = "area_terreno"
            elif col.startswith("area_construida"):
                COL_MAP[col] = "area_construida"
            elif col.startswith("base_de_calculo"):
                COL_MAP[col] = "valor_itbi"
            elif col.startswith("valor_venal_de_referencia_prop"):
                COL_MAP[col] = "valor_itbi"
            elif col.startswith("descricao_do_uso"):
                COL_MAP[col] = "descricao_uso"
            elif col.startswith("uso_") or col == "uso":
                COL_MAP[col] = "tipo_uso"
            elif col.startswith("tipo_de_financiamento") or col.startswith("tipo_financiamento"):
                COL_MAP[col] = "tipo_financiamento"
            elif col.startswith("acc_") or col == "acc":
                COL_MAP[col] = "acc_iptu"

    rename = {c: COL_MAP[c] for c in df.columns if c in COL_MAP}
    df = df.rename(columns=rename)
    df = df.loc[:, ~df.columns.duplicated(keep='first')]

    df["ano_referencia"] = ano
    df["mes_referencia"] = mes

    def limpar_numero(x):
        if x is None:
            return None
        s = str(x).strip()
        s = re.sub(r'[\s\t]', '', s)
        s = re.sub(r'[R$]', '', s)
        s = s.strip()
        if not s:
            return None
        tem_virgula = ',' in s
        num_pontos  = s.count('.')
        if tem_virgula and num_pontos >= 1:
            s = s.replace('.', '').replace(',', '.')
        elif tem_virgula and num_pontos == 0:
            s = s.replace(',', '.')
        elif not tem_virgula and num_pontos >= 2:
            s = s.replace('.', '')
        s = re.sub(r'[^\d.\-]', '', s)
        return s if s else None

    for col in ["valor_declarado", "valor_financiado", "valor_itbi",
                "area_terreno", "area_construida", "proporcao_transmitida"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col].apply(limpar_numero), errors="coerce")

    if "sql_terreno" in df.columns:
        def normalizar_sql(v):
            if not v:
                return None
            s = str(v).strip()
            try:
                if 'e' in s.lower() and '.' in s:
                    s = str(int(float(s)))
            except Exception:
                pass
            if s.endswith('.0'):
                s = s[:-2]
            if re.match(r'^\d{3}\.\d{3}\.\d{4}-\d$', s):
                return s
            digits = re.sub(r'[^\d]', '', s)
            if len(digits) == 11:
                return f"{digits[:3]}.{digits[3:6]}.{digits[6:10]}-{digits[10]}"
            return s if s not in ('nan', 'None', '') else None
        df["sql_terreno"] = df["sql_terreno"].apply(normalizar_sql)

    if "cep" in df.columns:
        def normalizar_cep(v):
            if not v:
                return None
            digits = ''.join(c for c in str(v) if c.isdigit())
            return digits.zfill(8) if digits else None
        df["cep"] = df["cep"].apply(normalizar_cep)

    for col in COLUNAS_FINAIS:
        if col not in df.columns:
            df[col] = None

    df = df[df["logradouro"].notna() | df["bairro"].notna() | df["valor_declarado"].notna()]
    df = df[df["logradouro"] != "Nome do Logradouro"]

    return df[COLUNAS_FINAIS]


# ──────────────────────────────────────────────
# PROCESSAR — XLSX → CSV → DataFrame
# A principal diferença: usa pandas para ler cada aba do XLSX
# e faz cache dos resultados em CSV. Reruns leem do CSV direto.
# ──────────────────────────────────────────────

def _df_raw_from_rows(rows: list, sheet_name: str) -> pd.DataFrame | None:
    """Recebe lista de linhas (de pd.read_excel header=None) e detecta header."""
    if len(rows) < 2:
        return None

    PALAVRAS = {"logradouro", "bairro", "cadastro", "natureza",
                "valor", "cep", "numero", "complemento"}

    header_idx = None
    for i, row in enumerate(rows[:15]):
        vals = {str(v).lower() for v in row if v is not None and str(v).strip()}
        if len(vals & PALAVRAS) >= 2:
            header_idx = i
            break

    if header_idx is None:
        n_cols = len(rows[0])
        raw_headers = list(COLS_POSICIONAIS[:n_cols])
        for j in range(len(raw_headers), n_cols):
            raw_headers.append(f"_col{j}")
        data_rows = rows
    else:
        raw_headers = [str(h).strip() if h is not None else f"_col{i}"
                       for i, h in enumerate(rows[header_idx])]
        data_rows = rows[header_idx + 1:]

    if not data_rows:
        return None

    # Garante nomes únicos
    seen: dict = {}
    headers = []
    for h in raw_headers:
        if h in seen:
            seen[h] += 1
            h = f"{h}_{seen[h]}"
        else:
            seen[h] = 0
        headers.append(h)

    df = pd.DataFrame(data_rows, columns=headers, dtype=object)
    df = df.where(df.notna(), None)
    df = df.dropna(how="all")
    return df if not df.empty else None


def processar_ano_csv(xlsx_path: Path, ano: int, forcar_csv: bool = False) -> pd.DataFrame:
    """
    Processa um arquivo XLSX ano a ano:
    1. Se XLSX não existir mas CSVs existirem → lê direto dos CSVs (sem abrir XLSX)
    2. Se XLSX existir → lê abas uma a uma, cacheia cada uma como CSV
    3. Normaliza e retorna DataFrame consolidado
    """
    # ── Modo CSV-only: sem XLSX na VM ─────────────────────────────────────
    if not xlsx_path.exists():
        csv_files = sorted(CSV_DIR.glob(f"itbi_{ano}_*.csv"))
        if not csv_files:
            print(f"   [ERRO] Nem XLSX nem CSVs para {ano}")
            return pd.DataFrame()
        frames_csv = []
        for csv_path in csv_files:
            sheet_name = csv_path.stem[len(f"itbi_{ano}_"):]
            mes = mes_da_aba(sheet_name)
            ano_sheet = ano_da_aba(sheet_name) or ano
            try:
                df_raw = pd.read_csv(csv_path, dtype=str, keep_default_na=False)
                df_raw = df_raw.replace({'': None, 'nan': None, 'None': None})
                df = normalizar_df(df_raw, ano_sheet, mes)
                if not df.empty:
                    frames_csv.append(df)
                    print(f"      [CSV] {sheet_name:<14} > {len(df):>7,} registros")
                del df_raw; gc.collect()
            except Exception as e:
                print(f"      [ERRO] {csv_path.name}: {e}")
        return pd.concat(frames_csv, ignore_index=True) if frames_csv else pd.DataFrame()

    # ── Modo normal: XLSX presente ────────────────────────────────────────
    from openpyxl import load_workbook

    # Lê só os nomes das abas — muito mais leve que carregar tudo
    wb_meta = load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet_names = [s for s in wb_meta.sheetnames
                   if not any(x in s.lower() for x in IGNORAR)]
    wb_meta.close()

    frames = []

    for sheet_name in sheet_names:
        mes = mes_da_aba(sheet_name)
        ano_sheet = ano_da_aba(sheet_name) or ano
        # Nome seguro para o arquivo CSV
        safe_name = re.sub(r'[^\w]', '_', sheet_name)
        csv_path  = CSV_DIR / f"itbi_{ano}_{safe_name}.csv"

        # ── Tenta ler do CSV em cache ─────────────────────
        if csv_path.exists() and not forcar_csv:
            try:
                df_raw_cached = pd.read_csv(csv_path, dtype=str, keep_default_na=False)
                df_raw_cached = df_raw_cached.replace({'': None, 'nan': None, 'None': None})
                df = normalizar_df(df_raw_cached, ano_sheet, mes)
                if not df.empty:
                    frames.append(df)
                    print(f"      📂 {sheet_name:<14} → {len(df):>7,} registros  (CSV cache)")
                    del df_raw_cached, df
                    gc.collect()
                    continue
            except Exception:
                pass  # CSV corrompido → refaz do XLSX

        # ── Lê do XLSX com pandas (1 aba por vez) ─────────
        try:
            df_xlsx = pd.read_excel(
                xlsx_path,
                sheet_name=sheet_name,
                header=None,       # detectamos o header manualmente
                dtype=str,         # tudo como string → menos uso de memória
                engine="openpyxl",
            )
            rows = [list(r) for r in df_xlsx.itertuples(index=False, name=None)]
            del df_xlsx
            gc.collect()

            df_raw = _df_raw_from_rows(rows, sheet_name)
            del rows
            gc.collect()

            if df_raw is None or df_raw.empty:
                continue

            # Salva CSV para reruns futuros
            try:
                df_raw.to_csv(csv_path, index=False)
            except Exception:
                pass  # falha no cache não é crítica

            df = normalizar_df(df_raw, ano_sheet, mes)
            del df_raw
            gc.collect()

            if df.empty:
                continue

            frames.append(df)
            mes_str = f"mês {mes}" if mes else "mês ?"
            print(f"      📅 {sheet_name:<14} → {len(df):>7,} registros  ({mes_str}) → CSV salvo")
            del df
            gc.collect()

        except Exception as e:
            import traceback
            print(f"      ⚠️  Aba '{sheet_name}' erro: {type(e).__name__}: {e}")
            print(traceback.format_exc())

    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()


# ──────────────────────────────────────────────
# BANCO — SALVAR
# ──────────────────────────────────────────────

def salvar_no_banco(df: pd.DataFrame, ano: int, hash_arquivo: str):
    conn = sqlite3.connect(DB_PATH)
    anos_df = sorted({int(a) for a in df["ano_referencia"].dropna().unique()})
    for _a in (anos_df or [ano]):
        conn.execute("DELETE FROM transacoes WHERE ano_referencia = ?", (_a,))
    df.to_sql("transacoes", conn, if_exists="append", index=False)
    conn.execute("""
        INSERT OR REPLACE INTO arquivos_processados (ano, hash, linhas, updated_at)
        VALUES (?, ?, ?, ?)
    """, (ano, hash_arquivo, len(df), datetime.now().isoformat()))
    conn.commit()
    conn.close()


def hash_em_banco(ano: int) -> str | None:
    if not DB_PATH.exists():
        return None
    conn = sqlite3.connect(DB_PATH)
    row = conn.execute(
        "SELECT hash FROM arquivos_processados WHERE ano = ?", (ano,)
    ).fetchone()
    conn.close()
    return row[0] if row else None


# ──────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────

def sincronizar(anos: list[int] = None, forcar: bool = False, limpar_csv: bool = False):
    init_db()
    anos_alvo = sorted(anos or ANOS_DEFAULT, reverse=True)
    total_geral = 0

    if limpar_csv:
        for f in CSV_DIR.glob("*.csv"):
            f.unlink()
        print("🗑️  CSVs em cache removidos.")

    from datetime import datetime as _dt
    ano_atual = _dt.now().year
    ano_anterior = ano_atual - 1

    for ano in anos_alvo:
        url = ARQUIVOS.get(ano)
        # Para o ano atual e anterior, sempre tenta pegar URL dinâmica da página oficial
        if ano == ano_atual or not url:
            url_din = buscar_url_dinamica(ano)
            if url_din:
                url = url_din
            elif not url:
                print(f"⚠️  Ano {ano} sem URL configurada e não encontrado na página.")
                continue

        xlsx_path = CACHE_DIR / f"itbi_{ano}.xlsx"

        # ── Prioridade 1: CSVs já existem → usa sem precisar do XLSX ─────
        csvs_existentes = list(CSV_DIR.glob(f"itbi_{ano}_*.csv"))
        if csvs_existentes and not forcar:
            print(f"   [CSV] [{ano}] {len(csvs_existentes)} CSVs em cache — processando direto...")
            df = processar_ano_csv(xlsx_path, ano, forcar_csv=False)
            if not df.empty:
                hash_ref = hash_em_banco(ano) or f"csv_{ano}"
                salvar_no_banco(df, ano, hash_ref)
                total_geral += len(df)
                print(f"   [OK] [{ano}] {len(df):,} registros salvos (do CSV).\n")
                del df; gc.collect()
            continue

        # ── Prioridade 2: XLSX local com hash igual → pula ───────────────
        if xlsx_path.exists() and not forcar:
            hash_salvo = hash_em_banco(ano)
            novo_hash  = file_hash(xlsx_path)
            if hash_salvo and novo_hash == hash_salvo:
                print(f"   [=] [{ano}] Arquivo nao mudou, pulando.\n")
                continue

        # ── Prioridade 3: Baixa do site ───────────────────────────────────
        if not xlsx_path.exists() or forcar:
            xlsx_baixado = download(ano, url)
            if xlsx_baixado is None:
                print(f"   [ERRO] [{ano}] Download falhou e sem CSVs em cache. Pulando.\n")
                continue
            xlsx_path = xlsx_baixado

        novo_hash = file_hash(xlsx_path)
        hash_salvo = hash_em_banco(ano)

        if not forcar and novo_hash == hash_salvo:
            print(f"   ↩️  [{ano}] Arquivo não mudou, pulando.\n")
            continue

        print(f"   📊 [{ano}] Convertendo XLSX → CSV e processando...")
        df = processar_ano_csv(xlsx_path, ano, forcar_csv=forcar)

        if df.empty:
            print(f"   ❌ [{ano}] Nenhum dado extraído.\n")
            continue

        salvar_no_banco(df, ano, novo_hash)
        total_geral += len(df)
        print(f"   ✅ [{ano}] {len(df):,} registros salvos.\n")
        del df; gc.collect()

    print(f"\n🎉 Pronto! Total inserido: {total_geral:,} registros.")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Scraper ITBI otimizado para CSV")
    parser.add_argument("--anos", nargs="+", type=int, help="Anos a processar (default: 2024-2026; aceita qualquer ano de 2006-2026)")
    parser.add_argument("--forcar", action="store_true", help="Reprocessa mesmo sem mudança de hash")
    parser.add_argument("--limpar-csv", action="store_true", dest="limpar_csv",
                        help="Remove CSVs em cache antes de processar")
    args = parser.parse_args()
    sincronizar(anos=args.anos, forcar=args.forcar, limpar_csv=args.limpar_csv)
