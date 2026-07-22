"""
exportar.py — Geração de Excel e PDF a partir dos dados ITBI-SP
"""
import io
import sqlite3
from pathlib import Path
from datetime import datetime
from typing import Optional

import pandas as pd

DB_PATH = Path(__file__).parent / "itbi.db"

# Colunas na ordem da planilha original do governo
COLUNAS = [
    ("sql_terreno",           "N° do Cadastro (SQL)"),
    ("logradouro",            "Nome do Logradouro"),
    ("numero",                "Número"),
    ("complemento",           "Complemento"),
    ("bairro",                "Bairro"),
    ("cep",                   "CEP"),
    ("natureza_transacao",    "Natureza de Transação"),
    ("valor_declarado",       "Valor de Transação (declarado pelo contribuinte)"),
    ("data_transacao",        "Data de Transação"),
    ("valor_venal_ref",       "Valor Venal de Referência"),
    ("proporcao_transmitida", "Proporção Transmitida (%)"),
    ("valor_itbi",            "Base de Cálculo adotada"),
    ("tipo_financiamento",    "Tipo de Financiamento"),
    ("valor_financiado",      "Valor Financiado"),
    ("cartorio_registro",     "Cartório de Registro"),
    ("matricula_imovel",      "Matrícula do Imóvel"),
    ("situacao_sql",          "Situação do SQL"),
    ("area_terreno",          "Área do Terreno (m2)"),
    ("testada",               "Testada (m)"),
    ("fracao_ideal",          "Fração Ideal"),
    ("area_construida",       "Área Construída (m2)"),
    ("tipo_uso",              "Uso (IPTU)"),
    ("descricao_uso",         "Descrição do uso (IPTU)"),
    ("padrao_iptu",           "Padrão (IPTU)"),
    ("acc_iptu",              "ACC (IPTU)"),
    ("ano_referencia",        "Ano Referência"),
    ("mes_referencia",        "Mês Referência"),
]

COLS_MONETARIAS = {"valor_declarado", "valor_financiado", "valor_itbi", "valor_venal_ref"}
COLS_NUMERICAS  = {"area_terreno", "area_construida", "testada", "proporcao_transmitida"}


# ──────────────────────────────────────────────
# CONSULTA
# ──────────────────────────────────────────────

def _multi_like(col: str, value: str) -> tuple[str, list]:
    """Suporte a múltiplos valores separados por vírgula → OR no SQL."""
    vals = [v.strip() for v in value.split(",") if v.strip()]
    if not vals:
        return "", []
    clauses = [f"UPPER({col}) LIKE UPPER(?)" for _ in vals]
    return f"({' OR '.join(clauses)})", [f"%{v}%" for v in vals]


def buscar(filtros: dict, ids: list = None, limit: int = 100_000,
           sort_col: str = "ano_referencia", sort_dir: str = "DESC") -> pd.DataFrame:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row

    if ids:
        ph = ','.join('?' * len(ids))
        rows = conn.execute(
            f"SELECT * FROM transacoes WHERE id IN ({ph}) ORDER BY {sort_col} {sort_dir}",
            ids
        ).fetchall()
    else:
        conds, params = [], []
        if filtros.get('logradouro'):
            cl, pr = _multi_like("logradouro", filtros['logradouro'])
            conds.append(cl); params.extend(pr)
        if filtros.get('numero'):
            conds.append("numero = ?")
            params.append(filtros['numero'].strip())
        if filtros.get('bairro'):
            cl, pr = _multi_like("bairro", filtros['bairro'])
            conds.append(cl); params.extend(pr)
        if filtros.get('cep'):
            cep_digits = ''.join(c for c in filtros['cep'] if c.isdigit()).zfill(8)
            conds.append("REPLACE(REPLACE(cep,'-',''),' ','') LIKE ?")
            params.append(f"{cep_digits}%")
        if filtros.get('ano_min'):
            conds.append("ano_referencia >= ?")
            params.append(int(filtros['ano_min']))
        if filtros.get('ano_max'):
            conds.append("ano_referencia <= ?")
            params.append(int(filtros['ano_max']))
        if filtros.get('valor_min'):
            conds.append("valor_declarado >= ?")
            params.append(float(filtros['valor_min']))
        if filtros.get('valor_max'):
            conds.append("valor_declarado <= ?")
            params.append(float(filtros['valor_max']))
        if filtros.get('natureza'):
            cl, pr = _multi_like("natureza_transacao", filtros['natureza'])
            conds.append(cl); params.extend(pr)

        where = ("WHERE " + " AND ".join(conds)) if conds else ""
        rows = conn.execute(
            f"SELECT * FROM transacoes {where} ORDER BY {sort_col} {sort_dir} LIMIT {limit}",
            params
        ).fetchall()

    conn.close()
    return pd.DataFrame([dict(r) for r in rows]) if rows else pd.DataFrame()


# ──────────────────────────────────────────────
# EXCEL
# ──────────────────────────────────────────────

def gerar_excel(df: pd.DataFrame, filtros: dict = None) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1

    wb = Workbook()

    # ── Aba 1: Dados ──────────────────────────
    ws = wb.active
    ws.title = "ITBI - Dados"

    AZUL_ESCURO = "1A2F6B"
    AZUL_CLARO  = "E8EDFF"
    CINZA_LINHA = "F5F7FF"

    side = Side(style='thin', color='C5CCE8')
    borda = Border(left=side, right=side, top=side, bottom=side)

    # Cabeçalho
    for ci, (_, nome) in enumerate(COLUNAS, 1):
        cell = ws.cell(row=1, column=ci, value=nome)
        cell.font      = Font(bold=True, color="FFFFFF", size=9, name="Calibri")
        cell.fill      = PatternFill("solid", fgColor=AZUL_ESCURO)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = borda
    ws.row_dimensions[1].height = 40

    # Dados
    col_keys = [c for c, _ in COLUNAS]
    for ri, row in enumerate(df.itertuples(index=False), 2):
        bg = CINZA_LINHA if ri % 2 == 0 else "FFFFFF"
        for ci, ck in enumerate(col_keys, 1):
            val = getattr(row, ck, None) if ck in df.columns else None
            # Converte NaN/None
            if pd.isna(val) if isinstance(val, float) else val is None:
                val = None
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.border    = borda
            cell.alignment = Alignment(vertical="center")
            if ck in COLS_MONETARIAS and val is not None:
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif ck in COLS_NUMERICAS and val is not None:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")

    # Larguras
    larguras = {
        "sql_terreno": 20, "logradouro": 32, "numero": 8, "complemento": 16,
        "bairro": 22, "cep": 12, "natureza_transacao": 35,
        "valor_declarado": 22, "data_transacao": 15,
        "valor_venal_ref": 22, "proporcao_transmitida": 16,
        "valor_itbi": 22, "tipo_financiamento": 22, "valor_financiado": 22,
        "cartorio_registro": 25, "matricula_imovel": 18,
        "situacao_sql": 15, "area_terreno": 16, "testada": 12,
        "fracao_ideal": 14, "area_construida": 16,
        "tipo_uso": 12, "descricao_uso": 25, "padrao_iptu": 14,
        "acc_iptu": 12, "ano_referencia": 10, "mes_referencia": 8,
    }
    for ci, (ck, _) in enumerate(COLUNAS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = larguras.get(ck, 14)

    ws.freeze_panes = "A2"

    # ── Aba 2: Resumo ─────────────────────────
    ws2 = wb.create_sheet("Resumo")

    def vbr(v):
        if v is None or (isinstance(v, float) and pd.isna(v)): return "—"
        return f"R$ {int(round(v)):,}".replace(",", ".")

    def estilo_resumo(cell, negrito=False, cor_fundo=None, cor_texto="000000"):
        cell.font = Font(bold=negrito, size=10, name="Calibri", color=cor_texto)
        if cor_fundo:
            cell.fill = PatternFill("solid", fgColor=cor_fundo)
        cell.border = borda
        cell.alignment = Alignment(vertical="center", indent=1)

    titulo = ws2.cell(1, 1, "Relatório ITBI · São Paulo")
    titulo.font = Font(bold=True, size=16, color="1A2F6B", name="Calibri")
    ws2.merge_cells("A1:C1")
    ws2.row_dimensions[1].height = 30

    ws2.cell(2, 1, f"Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}")
    ws2.merge_cells("A2:C2")

    linhas_resumo = [
        ("", "", ""),
        ("MÉTRICAS", "Valor", ""),
        ("Total de transações", f"{len(df):,}".replace(",", "."), ""),
        ("Volume total (R$)", vbr(df["valor_declarado"].sum() if "valor_declarado" in df.columns else 0), ""),
        ("Ticket médio (R$)", vbr(df["valor_declarado"].mean() if "valor_declarado" in df.columns else 0), ""),
        ("Maior transação (R$)", vbr(df["valor_declarado"].max() if "valor_declarado" in df.columns else 0), ""),
        ("Menor transação (R$)", vbr(df["valor_declarado"].min() if "valor_declarado" in df.columns else 0), ""),
        ("Base ITBI total (R$)", vbr(df["valor_itbi"].sum() if "valor_itbi" in df.columns else 0), ""),
    ]
    if filtros:
        linhas_resumo.append(("", "", ""))
        linhas_resumo.append(("FILTROS APLICADOS", "", ""))
        for k, v in filtros.items():
            if v: linhas_resumo.append((k, str(v), ""))

    for ri, (a, b, c) in enumerate(linhas_resumo, 4):
        ca = ws2.cell(ri, 1, a)
        cb = ws2.cell(ri, 2, b)
        negrito = a in ("MÉTRICAS", "FILTROS APLICADOS")
        bg = AZUL_ESCURO if negrito else ("FFFFFF" if ri % 2 == 0 else CINZA_LINHA)
        cor_tx = "FFFFFF" if negrito else "000000"
        estilo_resumo(ca, negrito, bg, cor_tx)
        estilo_resumo(cb, negrito, bg, cor_tx)

    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 28

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ──────────────────────────────────────────────
# PDF — estilo relatório profissional (light)
# ──────────────────────────────────────────────

# Paleta clara — cores normais, fundo branco, igual ao dashboard em modo claro
_BG       = '#ffffff'
_SURFACE  = '#f7f8fb'
_SURFACE2 = '#eef0f6'
_NAVY     = '#3a5bff'   # azul como "destaque principal"
_INK      = '#181b2c'
_INK2     = '#3a3e54'
_MU       = '#6b7280'
_LINE     = '#e2e5ee'
_BLUE     = '#3a5bff'
_GREEN    = '#159a5d'
_PURPLE   = '#7c4dff'
_AMBER    = '#c98a16'
_RED      = '#e23f7e'
_TEAL     = '#0ea5e9'

# Paleta de cores para gráficos (categorias, tons vivos sobre fundo branco)
_CHART_COLORS = ['#3a5bff','#7c4dff','#159a5d','#c98a16','#e23f7e',
                 '#0ea5e9','#f97316','#10b981','#8b5cf6','#ec4899']


def _fmt(v, mode='brl'):
    """Formata valor monetário ou numérico."""
    if v is None or (isinstance(v, float) and pd.isna(v)) or v == 0:
        return '—'
    n = float(v)
    if mode == 'brl':
        if n >= 1e9:  return f'R$ {n/1e9:.1f} Bi'.replace('.', ',')
        if n >= 1e6:  return f'R$ {n/1e6:.1f} Mi'.replace('.', ',')
        if n >= 1e3:  return f'R$ {n/1e3:.0f} k'.replace('.', ',')
        return f'R$ {int(round(n)):,}'.replace(',', '.')
    return f'{int(round(n)):,}'.replace(',', '.')


def _trunc(v, n):
    s = str(v or '').strip()
    return s[:n-1] + '…' if len(s) > n else s


def _draw_page(canvas, doc):
    """Cabeçalho e rodapé padrão em todas as páginas."""
    from reportlab.lib import colors
    W_PAGE, H_PAGE = doc.pagesize
    canvas.saveState()
    # Fundo total da página branco
    canvas.setFillColor(colors.HexColor(_BG))
    canvas.rect(0, 0, W_PAGE, H_PAGE, fill=1, stroke=0)
    # Barra de topo azul
    canvas.setFillColor(colors.HexColor(_BLUE))
    canvas.rect(0, H_PAGE - 28, W_PAGE, 28, fill=1, stroke=0)
    # Título no topo
    canvas.setFont('Helvetica-Bold', 9)
    canvas.setFillColor(colors.white)
    canvas.drawString(18, H_PAGE - 18, 'ITBI · São Paulo  —  Relatório de Transações Imobiliárias')
    canvas.setFont('Helvetica', 8)
    canvas.drawRightString(W_PAGE - 18, H_PAGE - 18,
                           f'Pág. {doc.page}  ·  {datetime.now().strftime("%d/%m/%Y")}')
    # Linha de rodapé
    canvas.setStrokeColor(colors.HexColor(_LINE))
    canvas.setLineWidth(0.5)
    canvas.line(18, 20, W_PAGE - 18, 20)
    canvas.setFont('Helvetica', 6.5)
    canvas.setFillColor(colors.HexColor(_MU))
    canvas.drawString(18, 8, 'Fonte: Secretaria Municipal de Financas de Sao Paulo · prefeitura.sp.gov.br/web/fazenda')
    canvas.restoreState()


def gerar_insights(df: pd.DataFrame, filtros: dict = None) -> str:
    """Gera análise textual automática baseada nos dados. Sem dependência de API externa."""
    if df.empty or len(df) < 2:
        return ""

    try:
        total = len(df)
        has_val = 'valor_declarado' in df.columns
        has_itbi = 'valor_itbi' in df.columns
        has_ano  = 'ano_referencia' in df.columns
        has_bairro = 'bairro' in df.columns

        vol  = df['valor_declarado'].sum()  if has_val else 0
        med  = df['valor_declarado'].mean() if has_val else 0
        maxi = df['valor_declarado'].max()  if has_val else 0
        itbi = df['valor_itbi'].sum()       if has_itbi else 0

        paras = []

        # ── Parágrafo 1: visão geral ──────────────────────────────────
        filtros_str = ""
        if filtros:
            partes = []
            if filtros.get('bairro'):    partes.append("bairro " + filtros['bairro'])
            if filtros.get('logradouro'):partes.append("logradouro " + filtros['logradouro'])
            if filtros.get('ano_min') or filtros.get('ano_max'):
                a0 = filtros.get('ano_min', ''); a1 = filtros.get('ano_max', '')
                if a0 and a1:   partes.append(f"periodo {a0}-{a1}")
                elif a0:        partes.append(f"a partir de {a0}")
                else:           partes.append(f"ate {a1}")
            if filtros.get('natureza'):  partes.append("natureza: " + filtros['natureza'])
            if partes:
                filtros_str = " com filtros aplicados (" + ', '.join(partes) + ")"

        p1 = (f"O relatório consolidou {total:,} transações imobiliárias registradas na Prefeitura de São Paulo"
              f"{filtros_str}, totalizando um volume transacionado de "
              f"R$ {vol/1e6:.1f} milhão{'ões' if vol/1e6>=2 else ''} e base de cálculo ITBI de "
              f"R$ {itbi/1e6:.1f} Mi. O ticket médio registrado foi de "
              f"R$ {med/1e3:.0f} mil, com a maior transação individual alcançando "
              f"R$ {maxi/1e3:.0f} mil.").replace(",", ".")
        paras.append(p1)

        # ── Parágrafo 2: tendência temporal ──────────────────────────
        if has_ano and has_val:
            por_ano = (df.groupby('ano_referencia')['valor_declarado']
                         .agg(n='count', vol='sum', med='mean')
                         .reset_index().sort_values('ano_referencia'))
            if len(por_ano) >= 3:
                anos_rec = por_ano.tail(3)
                n_vals = anos_rec['n'].tolist()
                v_vals = anos_rec['vol'].tolist()
                anos_l = anos_rec['ano_referencia'].astype(int).tolist()

                # tendência de volume
                if v_vals[-1] > v_vals[-2] * 1.05:
                    tend_vol = f"crescimento de {((v_vals[-1]/v_vals[-2])-1)*100:.0f}% em {anos_l[-1]} frente a {anos_l[-2]}"
                elif v_vals[-1] < v_vals[-2] * 0.95:
                    tend_vol = f"retração de {((1-v_vals[-1]/v_vals[-2]))*100:.0f}% em {anos_l[-1]} frente a {anos_l[-2]}"
                else:
                    tend_vol = f"estabilidade entre {anos_l[-2]} e {anos_l[-1]}"

                # ticket médio crescendo?
                med_vals = anos_rec['med'].tolist()
                if med_vals[-1] > med_vals[0] * 1.10:
                    tend_ticket = (f"O ticket médio apresentou trajetória de alta ao longo do período, "
                                   f"saindo de R$ {med_vals[0]/1e3:.0f} mil em {anos_l[0]} para "
                                   f"R$ {med_vals[-1]/1e3:.0f} mil em {anos_l[-1]}, "
                                   f"refletindo valorização real dos imóveis transacionados.")
                elif med_vals[-1] < med_vals[0] * 0.90:
                    tend_ticket = (f"O ticket médio recuou ao longo do período analisado, "
                                   f"de R$ {med_vals[0]/1e3:.0f} mil ({anos_l[0]}) para "
                                   f"R$ {med_vals[-1]/1e3:.0f} mil ({anos_l[-1]}), "
                                   f"possivelmente indicando maior participação de imóveis de menor valor.")
                else:
                    tend_ticket = (f"O ticket médio manteve-se relativamente estável no intervalo, "
                                   f"em torno de R$ {med_vals[-1]/1e3:.0f} mil.")

                p2 = (f"A análise temporal evidencia {tend_vol} em termos de volume financeiro. "
                      f"{tend_ticket}")
                paras.append(p2)

        # ── Parágrafo 3: concentração geográfica ─────────────────────
        if has_bairro:
            top_b = (df[df['bairro'].notna()].groupby('bairro')['id']
                       .count().nlargest(5))
            if len(top_b) > 0:
                top_nome  = top_b.index[0]
                top_pct   = top_b.iloc[0] / total * 100
                top3_pct  = top_b.head(3).sum() / total * 100
                top3_str  = ', '.join(top_b.head(3).index.tolist())

                p3 = (f"Em termos de distribuição geográfica, o bairro {top_nome} concentrou "
                      f"{top_pct:.1f}% das transações do período. Os três bairros mais ativos "
                      f"({top3_str}) responderam por {top3_pct:.1f}% do total, "
                      f"evidenciando {'forte' if top3_pct > 40 else 'moderada'} concentração territorial nas operações imobiliárias registradas.")
                paras.append(p3)

        # ── Parágrafo 4: alíquota efetiva ITBI ───────────────────────
        if has_val and has_itbi and vol > 0 and itbi > 0:
            aliq_ef = (itbi / vol) * 100
            p4 = (f"A alíquota efetiva média de ITBI apurada sobre o volume declarado foi de "
                  f"{aliq_ef:.2f}%, compatível com a legislação municipal vigente. "
                  f"A relação entre a base de cálculo adotada (R$ {itbi/1e6:.1f} Mi) e o volume "
                  f"declarado (R$ {vol/1e6:.1f} Mi) indica que "
                  f"{'a base venal de referência foi utilizada como piso de cálculo na maioria das operações' if aliq_ef < 3 else 'os valores declarados foram, em geral, adotados como base de cálculo'}.")
            paras.append(p4)

        return '\n\n'.join(paras)

    except Exception:
        return ""


def gerar_pdf(df: pd.DataFrame, filtros: dict = None) -> bytes:
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import matplotlib.ticker as mticker
    import numpy as np
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                     Table, TableStyle, Image as RLImage,
                                     HRFlowable, PageBreak)
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    C_BG    = colors.HexColor(_BG)
    C_SURF  = colors.HexColor(_SURFACE)
    C_SURF2 = colors.HexColor(_SURFACE2)
    C_NAVY  = colors.HexColor(_NAVY)
    C_INK   = colors.HexColor(_INK)
    C_INK2  = colors.HexColor(_INK2)
    C_MU    = colors.HexColor(_MU)
    C_LINE  = colors.HexColor(_LINE)
    C_BLUE  = colors.HexColor(_BLUE)
    C_GREEN = colors.HexColor(_GREEN)
    C_PURP  = colors.HexColor(_PURPLE)
    C_AMBER = colors.HexColor(_AMBER)
    C_TEAL  = colors.HexColor(_TEAL)

    PAGE = landscape(A4)
    W_PAGE, H_PAGE = PAGE
    MARGIN = 1.4 * cm
    W = W_PAGE - 2 * MARGIN

    def st(name, **kw):
        base = dict(fontName='Helvetica', fontSize=9, textColor=C_INK,
                    leading=13, spaceAfter=0, spaceBefore=0)
        base.update(kw)
        return ParagraphStyle(name, **base)

    s_title   = st('tit',  fontSize=20, fontName='Helvetica-Bold', textColor=C_NAVY, spaceAfter=2)
    s_sub     = st('sub',  fontSize=8,  textColor=C_MU, spaceAfter=10)
    s_section = st('sec',  fontSize=10, fontName='Helvetica-Bold', textColor=C_NAVY,
                   spaceBefore=12, spaceAfter=6)
    s_nota    = st('nota', fontSize=6.5, textColor=C_MU)
    s_cell    = st('cell', fontSize=7,  textColor=C_INK2, leading=10)
    s_cell_r  = st('cellr',fontSize=7,  textColor=C_INK2, leading=10, alignment=2)
    s_cell_b  = st('cellb',fontSize=7,  fontName='Helvetica-Bold', textColor=C_BLUE, leading=10, alignment=2)
    s_ai_body = st('aib',  fontSize=8,  textColor=C_INK2, leading=13)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=PAGE,
        leftMargin=MARGIN, rightMargin=MARGIN,
        topMargin=MARGIN + 0.8*cm,   # espaço para barra de topo
        bottomMargin=MARGIN,
        title='Relatorio ITBI - Sao Paulo',
    )

    elems = []
    total = len(df)
    vol  = df['valor_declarado'].sum()  if 'valor_declarado' in df.columns else 0
    med  = df['valor_declarado'].mean() if 'valor_declarado' in df.columns else 0
    maxi = df['valor_declarado'].max()  if 'valor_declarado' in df.columns else 0
    mini = df['valor_declarado'].min()  if 'valor_declarado' in df.columns else 0
    itbi = df['valor_itbi'].sum()       if 'valor_itbi'      in df.columns else 0

    # ── Cabeçalho ─────────────────────────────────────────────────────────
    # Título + data numa linha, filtros separados embaixo
    filtros_pares = [(k, v) for k, v in (filtros or {}).items() if v]

    # Linha do título com data alinhada à direita — usando tabela
    data_txt = datetime.now().strftime('%d/%m/%Y  %H:%M')
    hdr_tbl = Table(
        [[Paragraph('Relatório de Transações Imobiliárias', s_title),
          Paragraph(data_txt, st('dt', fontSize=8, textColor=C_MU, alignment=2))]],
        colWidths=[W * 0.72, W * 0.28]
    )
    hdr_tbl.setStyle(TableStyle([
        ('VALIGN',        (0,0), (-1,-1), 'BOTTOM'),
        ('LEFTPADDING',   (0,0), (-1,-1), 0),
        ('RIGHTPADDING',  (0,0), (-1,-1), 0),
        ('TOPPADDING',    (0,0), (-1,-1), 0),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
    ]))
    elems.append(hdr_tbl)
    elems.append(HRFlowable(width=W, thickness=2, color=C_NAVY, spaceAfter=6))

    # Filtros aplicados em chips coloridos
    if filtros_pares:
        nomes = {'ano_min': 'Ano de', 'ano_max': 'Ano ate', 'bairro': 'Bairro',
                 'logradouro': 'Logradouro', 'natureza': 'Natureza',
                 'valor_min': 'Valor min', 'valor_max': 'Valor max'}
        chips = '    '.join(
            f"[  {nomes.get(k, k)}: {v}  ]" for k, v in filtros_pares
        )
        elems.append(Paragraph(chips, st('chips', fontSize=7.5, textColor=C_BLUE, spaceAfter=10)))
    else:
        elems.append(Spacer(1, 10))

    # ── KPI cards (5 indicadores em linha) ────────────────────────────────
    kpi_data = [
        ('Transacoes',      _fmt(total, 'num'), '#3a5bff', '#f7f8fb'),
        ('Volume Total',    _fmt(vol),          '#7c4dff', '#f7f8fb'),
        ('Ticket Medio',    _fmt(med),          '#c98a16', '#f7f8fb'),
        ('Maior Transacao', _fmt(maxi),         '#159a5d', '#f7f8fb'),
        ('Base ITBI Total', _fmt(itbi),         '#e23f7e', '#f7f8fb'),
    ]
    cw_kpi = W / len(kpi_data)

    kpi_rows = []
    for label, value, accent, bg in kpi_data:
        c_acc = colors.HexColor(accent)
        c_bg  = colors.HexColor(bg)
        inner = Table(
            [[Paragraph(label, st('kl', fontSize=7, textColor=C_MU, alignment=1))],
             [Paragraph(value, st('kv', fontSize=14, fontName='Helvetica-Bold',
                                  textColor=c_acc, alignment=1))]],
            colWidths=[cw_kpi - 8]
        )
        inner.setStyle(TableStyle([
            ('BACKGROUND',    (0,0),(-1,-1), c_bg),
            ('TOPPADDING',    (0,0),(-1,-1), 10),
            ('BOTTOMPADDING', (0,0),(-1,-1), 10),
            ('LEFTPADDING',   (0,0),(-1,-1), 4),
            ('RIGHTPADDING',  (0,0),(-1,-1), 4),
            ('LINEABOVE',     (0,0),(-1, 0), 3, c_acc),
            ('ALIGN',         (0,0),(-1,-1), 'CENTER'),
        ]))
        kpi_rows.append(inner)

    kt = Table([kpi_rows], colWidths=[cw_kpi]*len(kpi_data))
    kt.setStyle(TableStyle([
        ('LEFTPADDING',  (0,0),(-1,-1), 4),
        ('RIGHTPADDING', (0,0),(-1,-1), 4),
        ('TOPPADDING',   (0,0),(-1,-1), 0),
        ('BOTTOMPADDING',(0,0),(-1,-1), 0),
    ]))
    elems.append(kt)
    elems.append(Spacer(1, 14))

    # ── Análise automática ────────────────────────────────────────────────
    insights = gerar_insights(df, filtros)
    if insights:
        elems.append(Paragraph('Analise do Periodo', s_section))
        paragraphs = [p.strip() for p in insights.split('\n') if p.strip()]
        ai_rows = [[Paragraph(p, s_ai_body)] for p in paragraphs]
        ai_tbl = Table(ai_rows, colWidths=[W])
        ai_tbl.setStyle(TableStyle([
            ('BACKGROUND',    (0,0),(-1,-1), C_SURF),
            ('TOPPADDING',    (0,0),(-1,-1), 5),
            ('BOTTOMPADDING', (0,0),(-1,-1), 5),
            ('LEFTPADDING',   (0,0),(-1,-1), 14),
            ('RIGHTPADDING',  (0,0),(-1,-1), 14),
            ('LINEBEFORE',    (0,0),(0,-1),  3, C_BLUE),
            ('LINEBELOW',     (0,-1),(-1,-1),0.5, C_LINE),
        ]))
        elems.append(ai_tbl)
        elems.append(Spacer(1, 14))

    # ── Gráficos ──────────────────────────────────────────────────────────
    if total > 0 and 'ano_referencia' in df.columns:
        elems.append(PageBreak())
        elems.append(Paragraph('Graficos e Indicadores', s_section))
        elems.append(Spacer(1, 6))

        plt.style.use('default')

        def _ax_style(ax, title=''):
            ax.set_facecolor('#ffffff')
            for sp in ax.spines.values():
                sp.set_color('#d8dce6'); sp.set_linewidth(0.6)
            ax.tick_params(colors='#6b7280', labelsize=8, length=0)
            ax.grid(color='#e2e5ee', linewidth=0.5, axis='y', zorder=0)
            ax.set_axisbelow(True)
            if title:
                ax.set_title(title, fontsize=10, fontweight='bold',
                             color=_BLUE, pad=10, loc='left')

        fmtK = mticker.FuncFormatter(lambda x, _: f'{int(x/1000)}k' if x>=1000 else str(int(x)))
        fmtM = mticker.FuncFormatter(lambda x, _:
            f'R${x/1e9:.1f}Bi' if x>=1e9 else
            f'R${x/1e6:.0f}Mi' if x>=1e6 else
            f'R${x/1e3:.0f}k')

        por_ano = (df.groupby('ano_referencia')
                     .agg(transacoes=('id','count'), volume=('valor_declarado','sum'),
                          ticket=('valor_declarado','mean'))
                     .reset_index().sort_values('ano_referencia'))
        anos_str = por_ano['ano_referencia'].astype(str).tolist()
        peak_t = por_ano['transacoes'].max()

        FIG_BG = '#ffffff'

        # ── Linha 1: Barras transações + Pizza natureza ──────────────────
        fig1, (axA, axB) = plt.subplots(1, 2, figsize=(16, 4.2),
                                         gridspec_kw={'width_ratios': [2, 1]})
        fig1.patch.set_facecolor(FIG_BG)

        _ax_style(axA, 'Transacoes por Ano')
        bar_colors = [_BLUE if v == peak_t else _SURFACE2 for v in por_ano['transacoes']]
        bars = axA.bar(anos_str, por_ano['transacoes'], color=bar_colors,
                       edgecolor='white', linewidth=0.5, width=0.65, zorder=3)
        for bar, val in zip(bars, por_ano['transacoes']):
            if val == peak_t:
                axA.text(bar.get_x()+bar.get_width()/2, val + peak_t*0.015,
                         f'{int(val):,}'.replace(',','.'),
                         ha='center', va='bottom', fontsize=8,
                         color=_BLUE, fontweight='bold')
        axA.tick_params(axis='x', rotation=45)
        axA.yaxis.set_major_formatter(fmtK)
        axA.set_facecolor(FIG_BG)

        _ax_style(axB, 'Natureza das Transacoes')
        if 'natureza_transacao' in df.columns:
            nat_g = (df['natureza_transacao'].str.replace(r'^\d+\.\s*', '', regex=True)
                       .str.strip().value_counts().head(6))
            pie_colors = _CHART_COLORS[:len(nat_g)]
            wedges, _, autotexts = axB.pie(
                nat_g.values, labels=None,
                colors=pie_colors, autopct='%1.0f%%', startangle=90,
                pctdistance=0.78,
                wedgeprops=dict(edgecolor='white', linewidth=2))
            for t in autotexts: t.set(fontsize=7.5, fontweight='bold', color='white')
            axB.legend(nat_g.index, loc='lower center', bbox_to_anchor=(0.5, -0.22),
                       fontsize=6.5, ncol=2, framealpha=0)
        axB.set_facecolor(FIG_BG)

        fig1.patch.set_facecolor(FIG_BG)
        plt.tight_layout(pad=1.5)
        ib1 = io.BytesIO()
        fig1.savefig(ib1, format='png', dpi=150, bbox_inches='tight', facecolor=FIG_BG)
        plt.close(fig1); ib1.seek(0)
        elems.append(RLImage(ib1, width=W, height=7.5*cm))
        elems.append(Spacer(1, 10))

        # ── Linha 2: Volume por ano + Top bairros ───────────────────────
        fig2, (axC, axD) = plt.subplots(1, 2, figsize=(16, 4.2))
        fig2.patch.set_facecolor(FIG_BG)

        _ax_style(axC, 'Volume Transacionado por Ano (R$)')
        axC.fill_between(range(len(anos_str)), por_ano['volume'],
                         alpha=0.15, color=_GREEN, zorder=2)
        axC.plot(range(len(anos_str)), por_ano['volume'], color=_GREEN, linewidth=2.5,
                 marker='o', markersize=5, markerfacecolor='white',
                 markeredgecolor=_GREEN, markeredgewidth=2, zorder=4)
        axC.set_xticks(range(len(anos_str))); axC.set_xticklabels(anos_str, rotation=45)
        axC.yaxis.set_major_formatter(fmtM)
        axC.set_facecolor(FIG_BG)
        idx_peak = por_ano['volume'].idxmax()
        loc_peak = por_ano.index.get_loc(idx_peak)
        axC.annotate(_fmt(por_ano.loc[idx_peak, 'volume']),
                     xy=(loc_peak, por_ano.loc[idx_peak, 'volume']),
                     xytext=(0, 12), textcoords='offset points',
                     ha='center', fontsize=7.5, color=_GREEN, fontweight='bold',
                     arrowprops=dict(arrowstyle='-', color=_GREEN, lw=0.8))

        _ax_style(axD, 'Top 10 Bairros por Transacoes')
        if 'bairro' in df.columns:
            top_b = (df[df['bairro'].notna()].groupby('bairro')
                       .agg(n=('id','count')).nlargest(10,'n').sort_values('n'))
            bar_h_colors = [_CHART_COLORS[i % len(_CHART_COLORS)] for i in range(len(top_b))]
            axD.barh(top_b.index, top_b['n'], color=bar_h_colors,
                     edgecolor='white', linewidth=0.5, height=0.65, zorder=3)
            axD.xaxis.set_major_formatter(fmtK)
            axD.tick_params(axis='y', labelsize=7.5)
            axD.grid(axis='x', color=_LINE, linewidth=0.5)
            axD.grid(axis='y', visible=False)
        axD.set_facecolor(FIG_BG)

        plt.tight_layout(pad=1.5)
        ib2 = io.BytesIO()
        fig2.savefig(ib2, format='png', dpi=150, bbox_inches='tight', facecolor=FIG_BG)
        plt.close(fig2); ib2.seek(0)
        elems.append(RLImage(ib2, width=W, height=7.5*cm))
        elems.append(Spacer(1, 10))

        # ── Linha 3: Ticket médio + Distribuição de valores ─────────────
        fig3, (axE, axF) = plt.subplots(1, 2, figsize=(16, 3.8))
        fig3.patch.set_facecolor(FIG_BG)

        _ax_style(axE, 'Ticket Medio por Ano (R$)')
        axE.fill_between(range(len(anos_str)), por_ano['ticket'],
                         alpha=0.12, color=_PURPLE, zorder=2)
        axE.plot(range(len(anos_str)), por_ano['ticket'], color=_PURPLE, linewidth=2.5,
                 marker='s', markersize=4, markerfacecolor='white',
                 markeredgecolor=_PURPLE, markeredgewidth=2, zorder=4)
        axE.set_xticks(range(len(anos_str))); axE.set_xticklabels(anos_str, rotation=45)
        axE.yaxis.set_major_formatter(fmtM)
        axE.set_facecolor(FIG_BG)

        _ax_style(axF, 'Distribuicao de Valores Declarados')
        vals = df['valor_declarado'].dropna()
        vals = vals[vals > 0]
        if len(vals) > 0:
            axF.hist(np.log10(vals.clip(upper=vals.quantile(0.99))),
                     bins=40, color=_TEAL, alpha=0.85, edgecolor='white',
                     linewidth=0.4, zorder=3)
            axF.xaxis.set_major_formatter(
                mticker.FuncFormatter(lambda x, _:
                    f'R${10**x/1e6:.0f}Mi' if x >= 6 else
                    f'R${10**x/1e3:.0f}k'  if x >= 3 else f'R${10**x:.0f}'))
        axF.set_ylabel('Frequencia', fontsize=8, color=_MU)
        axF.set_facecolor(FIG_BG)

        plt.tight_layout(pad=1.5)
        ib3 = io.BytesIO()
        fig3.savefig(ib3, format='png', dpi=150, bbox_inches='tight', facecolor=FIG_BG)
        plt.close(fig3); ib3.seek(0)
        elems.append(RLImage(ib3, width=W, height=6.8*cm))
        elems.append(Spacer(1, 10))

        # ── Linha 4: Sazonalidade mensal + Distribuição por faixa ───────
        fig4, (axG, axH) = plt.subplots(1, 2, figsize=(16, 4.0))
        fig4.patch.set_facecolor(FIG_BG)

        _ax_style(axG, 'Sazonalidade Mensal (ultimos 3 anos)')
        MESES_STR = ['Jan','Fev','Mar','Abr','Mai','Jun','Jul','Ago','Set','Out','Nov','Dez']
        if 'mes_referencia' in df.columns and 'ano_referencia' in df.columns:
            df_mes = df.dropna(subset=['mes_referencia', 'ano_referencia'])
            df_mes = df_mes[df_mes['mes_referencia'].between(1, 12)]
            ultimos_anos = sorted(df_mes['ano_referencia'].unique())[-3:]
            SAZ_CORES = [_BLUE, _GREEN, _AMBER]
            for i, ano in enumerate(ultimos_anos):
                sub = df_mes[df_mes['ano_referencia'] == ano]
                contagem = sub.groupby('mes_referencia').size()
                vals_mes = [contagem.get(m, np.nan) for m in range(1, 13)]
                xs = [j for j, v in enumerate(vals_mes) if not np.isnan(v)]
                ys = [v for v in vals_mes if not np.isnan(v)]
                if xs:
                    axG.plot(xs, ys, color=SAZ_CORES[i % 3], linewidth=2,
                             marker='o', markersize=4, markerfacecolor='white',
                             markeredgecolor=SAZ_CORES[i % 3], markeredgewidth=1.5,
                             label=str(int(ano)), zorder=4)
                    axG.fill_between(xs, ys, alpha=0.08, color=SAZ_CORES[i % 3], zorder=2)
            axG.set_xticks(range(12))
            axG.set_xticklabels(MESES_STR, fontsize=8)
            axG.yaxis.set_major_formatter(fmtK)
            axG.legend(fontsize=8, framealpha=0, loc='upper right')
        axG.set_facecolor(FIG_BG)

        _ax_style(axH, 'Distribuicao por Faixa de Valor')
        if 'valor_declarado' in df.columns:
            faixas_labels = ['Ate\nR$300k', 'R$300k\n600k', 'R$600k\n1M', 'R$1M\n2M', 'Acima\nR$2M']
            faixas_bins = [0, 300_000, 600_000, 1_000_000, 2_000_000, float('inf')]
            vd = df['valor_declarado'].dropna()
            vd = vd[vd > 0]
            contagens_fx = [((vd >= faixas_bins[j]) & (vd < faixas_bins[j+1])).sum()
                             for j in range(5)]
            max_fx = max(contagens_fx) if contagens_fx else 1
            bar_cols_fx = [_BLUE if c == max_fx else _SURFACE2 for c in contagens_fx]
            brs = axH.bar(range(5), contagens_fx, color=bar_cols_fx,
                          edgecolor='white', linewidth=0.5, width=0.6, zorder=3)
            for bar, val in zip(brs, contagens_fx):
                if val > 0:
                    axH.text(bar.get_x()+bar.get_width()/2, val + max_fx*0.01,
                             fmtK(val, None), ha='center', va='bottom',
                             fontsize=7.5, color=_MU)
            axH.set_xticks(range(5))
            axH.set_xticklabels(faixas_labels, fontsize=8)
            axH.yaxis.set_major_formatter(fmtK)
        axH.set_facecolor(FIG_BG)

        plt.tight_layout(pad=1.5)
        ib4 = io.BytesIO()
        fig4.savefig(ib4, format='png', dpi=150, bbox_inches='tight', facecolor=FIG_BG)
        plt.close(fig4); ib4.seek(0)
        elems.append(RLImage(ib4, width=W, height=7.2*cm))
        elems.append(Spacer(1, 6))

    # ── Tabela de registros ───────────────────────────────────────────────
    elems.append(PageBreak())
    MAX_ROWS = 500
    label_rows = f'{total:,} transacoes'.replace(',', '.')
    if total > MAX_ROWS:
        label_rows += f' · exibindo os primeiros {MAX_ROWS}'
    elems.append(Paragraph(f'Registros · {label_rows}', s_section))

    HD = ['Ano/Mes', 'Data', 'Logradouro', 'N', 'Bairro',
          'Valor Declarado', 'Base ITBI', 'Area Constr.', 'Natureza']
    CW = [1.9*cm, 2.2*cm, 6.0*cm, 1.1*cm, 4.0*cm,
          3.6*cm, 3.2*cm, 2.0*cm, 4.7*cm]

    s_hd = st('hd', fontSize=7, fontName='Helvetica-Bold', textColor=colors.white, alignment=1)

    rows_pdf = [[Paragraph(h, s_hd) for h in HD]]
    for _, r in df.head(MAX_ROWS).iterrows():
        am = str(int(r.get('ano_referencia') or 0))
        if r.get('mes_referencia'):
            am += f"/{int(r['mes_referencia']):02d}"
        nat = _trunc(str(r.get('natureza_transacao') or '').split('.', 1)[-1].strip(), 32)
        rows_pdf.append([
            Paragraph(am,                                    s_cell),
            Paragraph(_trunc(r.get('data_transacao',''),12), s_cell),
            Paragraph(_trunc(r.get('logradouro',''), 34),    s_cell),
            Paragraph(_trunc(r.get('numero',''), 6),         s_cell),
            Paragraph(_trunc(r.get('bairro',''), 24),        s_cell),
            Paragraph(_fmt(r.get('valor_declarado')),        s_cell_b),
            Paragraph(_fmt(r.get('valor_itbi')),             s_cell_r),
            Paragraph(str(r.get('area_construida','') or '—'), s_cell_r),
            Paragraph(nat,                                   s_cell),
        ])

    if total > MAX_ROWS:
        elems.append(Spacer(1, 4))
        elems.append(Paragraph(
            f'... e mais {total-MAX_ROWS:,} registros nao exibidos.'.replace(',', '.'), s_nota))

    t = Table(rows_pdf, colWidths=CW, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND',    (0,0), (-1,0),  C_NAVY),
        ('LINEBELOW',     (0,0), (-1,0),  2,    C_BLUE),
        ('ROWBACKGROUNDS',(0,1), (-1,-1), [C_SURF, C_SURF2]),
        ('GRID',          (0,0), (-1,-1), 0.3,  C_LINE),
        ('TOPPADDING',    (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('LEFTPADDING',   (0,0), (-1,-1), 5),
        ('RIGHTPADDING',  (0,0), (-1,-1), 5),
        ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
    ]))
    elems.append(t)

    # ── Fichas completas por transação ────────────────────────────────────
    MAX_FICHAS = 50
    fichas_df = df.head(MAX_FICHAS)

    if not fichas_df.empty:
        elems.append(PageBreak())
        label_fichas = f'Fichas Detalhadas · primeiras {len(fichas_df)} transações'
        if total > MAX_FICHAS:
            label_fichas += f' de {total:,}'.replace(',', '.')
        elems.append(Paragraph(label_fichas, s_section))
        elems.append(Spacer(1, 6))

        FICHA_W = (W - 10) / 2   # 2 fichas por página, com gap de 10 pt

        def _build_ficha(r):
            """Monta a ficha completa de uma transação como Table."""

            def _v(key, fmt=None):
                v = r.get(key)
                if v is None:
                    return '—'
                if isinstance(v, float) and pd.isna(v):
                    return '—'
                s = str(v).strip()
                if not s or s in ('nan', 'None', 'none'):
                    return '—'
                if fmt == 'brl':
                    return _fmt(v)
                if fmt == 'pct':
                    try:
                        return f'{float(v):.1f}%'
                    except Exception:
                        return s
                if fmt == 'm2':
                    try:
                        f = float(v)
                        return f'{f:,.2f} m²'.replace(',', '.')
                    except Exception:
                        return s
                return s

            LW = FICHA_W * 0.36
            VW = FICHA_W * 0.64 - 4

            s_gh  = ParagraphStyle('fgh',  fontName='Helvetica-Bold', fontSize=6.5,
                                    textColor=colors.white, leading=10)
            s_lbl = ParagraphStyle('flbl', fontName='Helvetica', fontSize=6,
                                    textColor=C_MU, leading=9)
            s_val = ParagraphStyle('fval', fontName='Helvetica-Bold', fontSize=7,
                                    textColor=C_INK, leading=10)
            s_fhd = ParagraphStyle('ffhd', fontName='Helvetica-Bold', fontSize=8,
                                    textColor=C_NAVY, leading=12)

            rows_f  = []
            stls_f  = []
            is_data = []

            def add_hdr(txt, c_hex):
                i = len(rows_f)
                rows_f.append([Paragraph(txt, s_gh), ''])
                is_data.append(False)
                stls_f.extend([
                    ('SPAN',          (0, i), (1, i)),
                    ('BACKGROUND',    (0, i), (1, i), colors.HexColor(c_hex)),
                    ('TOPPADDING',    (0, i), (1, i), 3),
                    ('BOTTOMPADDING', (0, i), (1, i), 3),
                    ('LEFTPADDING',   (0, i), (1, i), 6),
                    ('RIGHTPADDING',  (0, i), (1, i), 4),
                ])

            def add_row(lbl, val):
                i = len(rows_f)
                rows_f.append([Paragraph(lbl, s_lbl), Paragraph(val, s_val)])
                is_data.append(True)
                stls_f.extend([
                    ('TOPPADDING',    (0, i), (1, i), 2),
                    ('BOTTOMPADDING', (0, i), (1, i), 2),
                    ('LEFTPADDING',   (0, i), (0, i), 6),
                    ('RIGHTPADDING',  (0, i), (0, i), 3),
                    ('LEFTPADDING',   (1, i), (1, i), 4),
                    ('RIGHTPADDING',  (1, i), (1, i), 4),
                ])

            # ── Cabeçalho do imóvel ───────────────────────
            addr = _v('logradouro')
            if _v('numero') != '—':
                addr += f", {_v('numero')}"
            if _v('complemento') != '—':
                addr += f" {_v('complemento')}"
            i = len(rows_f)
            rows_f.append([Paragraph(addr, s_fhd), ''])
            is_data.append(False)
            stls_f += [
                ('SPAN',          (0, i), (1, i)),
                ('BACKGROUND',    (0, i), (1, i), C_SURF2),
                ('TOPPADDING',    (0, i), (1, i), 6),
                ('BOTTOMPADDING', (0, i), (1, i), 6),
                ('LEFTPADDING',   (0, i), (1, i), 6),
            ]

            # ── Transação ─────────────────────────────────
            add_hdr('TRANSAÇÃO', _NAVY)
            add_row('Data',       _v('data_transacao'))
            add_row('Ano / Mês',  f"{_v('ano_referencia')} / {_v('mes_referencia')}")
            add_row('Natureza',   _trunc(_v('natureza_transacao'), 44))
            add_row('Cartório',   _trunc(_v('cartorio_registro'), 30))
            add_row('Matrícula',  _v('matricula_imovel'))

            # ── Valores ───────────────────────────────────
            add_hdr('VALORES', _BLUE)
            add_row('Valor Declarado',    _v('valor_declarado',    'brl'))
            add_row('Base ITBI',          _v('valor_itbi',         'brl'))
            add_row('Valor Venal Ref.',   _v('valor_venal_ref',    'brl'))
            add_row('Valor Financiado',   _v('valor_financiado',   'brl'))
            add_row('Proporção',          _v('proporcao_transmitida', 'pct'))
            add_row('Tipo Financiamento', _v('tipo_financiamento'))

            # ── Localização ───────────────────────────────
            add_hdr('LOCALIZAÇÃO', _TEAL)
            add_row('Bairro',      _v('bairro'))
            add_row('CEP',         _v('cep'))
            add_row('SQL Terreno', _v('sql_terreno'))

            # ── Imóvel ────────────────────────────────────
            add_hdr('IMÓVEL', _PURPLE)
            add_row('Área Terreno',    _v('area_terreno',    'm2'))
            add_row('Área Construída', _v('area_construida', 'm2'))
            add_row('Testada',         (_v('testada') + ' m') if _v('testada') != '—' else '—')
            add_row('Fração Ideal',    _v('fracao_ideal'))

            # ── IPTU ──────────────────────────────────────
            add_hdr('IPTU', _AMBER)
            uso = _v('descricao_uso')
            if uso == '—':
                uso = _v('tipo_uso')
            add_row('Uso',         uso)
            add_row('Padrão',      _v('padrao_iptu'))
            add_row('ACC',         _v('acc_iptu'))
            add_row('Situação SQL', _v('situacao_sql'))

            # ── Estilos finais ────────────────────────────
            data_count = 0
            for i2, flag in enumerate(is_data):
                if flag:
                    bg = C_SURF if data_count % 2 == 0 else C_BG
                    stls_f.append(('BACKGROUND', (0, i2), (1, i2), bg))
                    data_count += 1

            stls_f += [
                ('GRID',   (0, 0), (-1, -1), 0.3, C_LINE),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('BOX',    (0, 0), (-1, -1), 1,   colors.HexColor(_LINE)),
            ]

            t_ficha = Table(rows_f, colWidths=[LW, VW])
            t_ficha.setStyle(TableStyle(stls_f))
            return t_ficha

        # Emparelha 2 fichas por linha
        fichas_list = [_build_ficha(r) for _, r in fichas_df.iterrows()]

        for i in range(0, len(fichas_list), 2):
            pair = fichas_list[i:i+2]
            if len(pair) == 1:
                # padding vazio para a segunda coluna
                pair.append(Spacer(FICHA_W, 1))
            row_tbl = Table([pair], colWidths=[FICHA_W, FICHA_W],
                            hAlign='LEFT', spaceAfter=8)
            row_tbl.setStyle(TableStyle([
                ('LEFTPADDING',  (0, 0), (-1, -1), 0),
                ('RIGHTPADDING', (0, 0), (-1, -1), 5),
                ('TOPPADDING',   (0, 0), (-1, -1), 0),
                ('BOTTOMPADDING',(0, 0), (-1, -1), 0),
                ('VALIGN',       (0, 0), (-1, -1), 'TOP'),
            ]))
            elems.append(row_tbl)

        if total > MAX_FICHAS:
            elems.append(Spacer(1, 6))
            elems.append(Paragraph(
                f'... e mais {total - MAX_FICHAS:,} transações não exibidas nas fichas.'.replace(',', '.'),
                s_nota))

    doc.build(elems, onFirstPage=_draw_page, onLaterPages=_draw_page)
    buf.seek(0)
    return buf.getvalue()
