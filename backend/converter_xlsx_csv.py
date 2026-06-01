"""
converter_xlsx_csv.py — Converte XLSXs locais (2020–2026) para CSV por aba.

Roda NO PC LOCAL (com mais RAM) antes de subir para a VM.
Os CSVs ficam em backend/cache/csv/ — pasta separada dos XLSXs originais.

Uso:
  python converter_xlsx_csv.py           # converte todos 2020–2026
  python converter_xlsx_csv.py --ano 2025
  python converter_xlsx_csv.py --forcar  # reconverte mesmo se CSV já existir

Rollback (se quiser desfazer):
  rm -rf backend/cache/csv/   (ou apaga a pasta manualmente)
  A VM continua com o scraper.py original sem CSVs.
"""

import re
import gc
import sys
import argparse
import pandas as pd
from pathlib import Path

CACHE_DIR = Path(__file__).parent / "cache"
CSV_DIR   = Path(__file__).parent / "cache" / "csv"
CSV_DIR.mkdir(parents=True, exist_ok=True)

ANOS_ALVO = [2020, 2021, 2022, 2023, 2024, 2025, 2026]

IGNORAR = {"legenda", "sumario", "sumário", "instrucao", "leia",
           "orientacao", "explicac"}

COLS_POSICIONAIS = [
    "N° do Cadastro (SQL)", "Nome do Logradouro", "Número",
    "Complemento", "Bairro", "Referência", "CEP",
    "Natureza de Transação",
    "Valor de Transação (declarado pelo contribuinte)",
    "Data de Transação", "Valor Venal de Referência",
    "Proporção Transmitida (%)",
    "Valor Venal de Referência Proporcional",
    "Tipo de Financiamento", "Valor Financiado",
    "Cartório de Registro", "Matrícula do Imóvel",
    "Situação do SQL", "Área do Terreno (m2)", "Testada (m)",
    "Fração Ideal", "Área Construída (m2)", "Uso (IPTU)",
    "Descrição do Uso (IPTU)", "Padrão (IPTU)", "ACC (IPTU)",
]

PALAVRAS_HEADER = {"logradouro", "bairro", "cadastro", "natureza",
                   "valor", "cep", "numero", "complemento"}


def _nome_csv(ano: int, sheet_name: str) -> Path:
    safe = re.sub(r'[^\w]', '_', sheet_name).strip('_')
    return CSV_DIR / f"itbi_{ano}_{safe}.csv"


def _detectar_header(df_raw: pd.DataFrame):
    """
    Recebe DataFrame sem cabeçalho (header=None).
    Retorna (header_idx, data_start_idx) ou (None, 0) se não encontrar.
    """
    rows = df_raw.values.tolist()
    for i, row in enumerate(rows[:15]):
        vals = {str(v).lower() for v in row if v is not None and str(v).strip()}
        if len(vals & PALAVRAS_HEADER) >= 2:
            return i, i + 1
    return None, 0


def converter_xlsx(ano: int, forcar: bool = False) -> int:
    xlsx_path = CACHE_DIR / f"itbi_{ano}.xlsx"
    if not xlsx_path.exists():
        print(f"  ⚠️  [{ano}] XLSX não encontrado em {xlsx_path}")
        return 0

    print(f"\n[XLSX] [{ano}] {xlsx_path.name} ({xlsx_path.stat().st_size / 1024 / 1024:.1f} MB)")

    # Lista abas sem carregar dados
    from openpyxl import load_workbook
    wb_meta = load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet_names = [s for s in wb_meta.sheetnames
                   if not any(x in s.lower() for x in IGNORAR)]
    wb_meta.close()

    csvs_gerados = 0

    for sheet_name in sheet_names:
        csv_path = _nome_csv(ano, sheet_name)

        if csv_path.exists() and not forcar:
            print(f"  [OK] {sheet_name:<18} >já existe ({csv_path.name})")
            csvs_gerados += 1
            continue

        try:
            # Lê aba com pandas — dtype=str para não inferir tipos e usar menos RAM
            df_raw = pd.read_excel(
                xlsx_path,
                sheet_name=sheet_name,
                header=None,
                dtype=str,
                engine="openpyxl",
            )
            df_raw = df_raw.dropna(how="all")

            if df_raw.empty:
                print(f"  [-] {sheet_name:<18} >vazia, pulada")
                continue

            header_idx, data_start = _detectar_header(df_raw)

            if header_idx is not None:
                # Usa linha de header encontrada
                headers = [str(v).strip() if v is not None else f"_col{i}"
                           for i, v in enumerate(df_raw.iloc[header_idx])]
                df_data = df_raw.iloc[data_start:].copy()
            else:
                # Sem header — usa colunas posicionais conhecidas
                n = len(df_raw.columns)
                headers = list(COLS_POSICIONAIS[:n])
                for j in range(len(headers), n):
                    headers.append(f"_col{j}")
                df_data = df_raw.copy()

            # Garante nomes únicos
            seen: dict = {}
            unique_headers = []
            for h in headers:
                if h in seen:
                    seen[h] += 1
                    h = f"{h}_{seen[h]}"
                else:
                    seen[h] = 0
                unique_headers.append(h)

            df_data.columns = unique_headers
            df_data = df_data.dropna(how="all")

            if df_data.empty:
                print(f"  [-] {sheet_name:<18} >sem dados após limpeza")
                del df_raw, df_data; gc.collect()
                continue

            # Salva CSV
            df_data.to_csv(csv_path, index=False, encoding="utf-8")
            print(f"  [CSV] {sheet_name:<18} >{len(df_data):>7,} linhas  >{csv_path.name}")
            csvs_gerados += 1

            del df_raw, df_data
            gc.collect()

        except Exception as e:
            print(f"  [ERRO] {sheet_name:<18} >ERRO: {e}")
            import traceback; traceback.print_exc()

    return csvs_gerados


def main():
    parser = argparse.ArgumentParser(description="Converte XLSX >CSV por aba")
    parser.add_argument("--ano",    type=int, help="Converter só este ano")
    parser.add_argument("--forcar", action="store_true", help="Reconverte mesmo se CSV já existe")
    args = parser.parse_args()

    anos = [args.ano] if args.ano else ANOS_ALVO
    total_csvs = 0

    print(f"Saida: {CSV_DIR}")
    print(f"Anos: {anos}\n")

    for ano in sorted(anos, reverse=True):
        n = converter_xlsx(ano, forcar=args.forcar)
        total_csvs += n

    print(f"\nPRONTO! Pronto! {total_csvs} arquivos CSV em {CSV_DIR}")
    print(f"\nPróximo passo — suba os CSVs para a VM:")
    print(f'  scp -i "$env:USERPROFILE\\.ssh\\itbi.key" -r {CSV_DIR} ubuntu@137.131.160.254:~/backend/cache/')


if __name__ == "__main__":
    main()
